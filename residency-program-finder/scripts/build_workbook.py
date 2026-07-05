#!/usr/bin/env python3
"""Build a sortable Excel workbook from a verified residency-program list.

Usage:
    python3 build_workbook.py programs.json --config config.json --out "list.xlsx"

See scripts/README.md for the programs.json / config.json schemas. Requires openpyxl.
"""
import argparse, json, sys

def load(p):
    with open(p) as f: return json.load(f)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("programs")
    ap.add_argument("--config", default=None)
    ap.add_argument("--out", default="program_list.xlsx")
    a = ap.parse_args()
    progs = load(a.programs)
    cfg = load(a.config) if a.config else {}

    applicant = cfg.get("applicant", "Applicant")
    specialty = cfg.get("specialty", "Internal Medicine")
    cycle = cfg.get("cycle", "")
    affinity = cfg.get("affinity_label", "Affinity-group")
    same_school = cfg.get("same_school_label", "own-school")
    gold_n = cfg.get("gold_count"); silver_n = cfg.get("silver_count")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    NAVY = "1F3A5F"
    thin = Side(style="thin", color="D0D0D0"); bd = Border(thin, thin, thin, thin)
    hf = PatternFill("solid", fgColor=NAVY); hfont = Font(color="FFFFFF", bold=True, size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    GREEN = PatternFill("solid", fgColor="D6EAD6"); YEL = PatternFill("solid", fgColor="FBF3D0")
    REACHF = PatternFill("solid", fgColor="E5D6F0"); BLUE = PatternFill("solid", fgColor="CFE2F3")
    GOLDF = PatternFill("solid", fgColor="F4E5B8"); SILF = PatternFill("solid", fgColor="E6E9ED")
    LIVE = PatternFill("solid", fgColor="E4F0EF")
    tfill = {"Reach": REACHF, "Target": GREEN, "Safety": YEL}

    def g(o, k, d=""):
        v = o.get(k)
        return v if v not in (None, "") else d

    sO = {"Gold": 0, "Silver": 1, "": 2}
    tO = {"Reach": 0, "Target": 1, "Safety": 2}
    progs = sorted(progs, key=lambda o: (0 if g(o, "signal") else 1,
                                         sO.get(g(o, "signal"), 2),
                                         tO.get(g(o, "tier"), 1),
                                         str(g(o, "name"))))

    wb = openpyxl.Workbook()

    ws0 = wb.active; ws0.title = "READ ME"
    ws0.column_dimensions["A"].width = 3; ws0.column_dimensions["B"].width = 118
    ng = sum(1 for o in progs if g(o, "signal") == "Gold")
    ns = sum(1 for o in progs if g(o, "signal") == "Silver")
    golds = " · ".join(o["name"] for o in progs if g(o, "signal") == "Gold")
    silvers = " · ".join(o["name"] for o in progs if g(o, "signal") == "Silver")
    nlive = sum(1 for o in progs if o.get("verified"))
    from collections import Counter
    tc = Counter(g(o, "tier", "Target") for o in progs)
    lines = [
        ("H", f"{specialty} residency program list — {applicant}" + (f" ({cycle})" if cycle else "")),
        ("", ""),
        ("S", "How this list was built"),
        ("T", f"{len(progs)} programs, ranked by the applicant's priorities. {nlive} were LIVE-verified on the "
              f"program's own current-resident roster / visa page (the authoritative source); the rest are "
              f"aggregator-sourced and labeled as such. Tiers: {tc.get('Reach',0)} reach · {tc.get('Target',0)} "
              f"target · {tc.get('Safety',0)} safety."),
        ("S", "Columns"),
        ("T", f"'Non-US IMG' = share of visa-requiring international grads in the program (not '% IMG'). "
              f"'{affinity} residents' = roster-confirmed count where the site publishes medical schools, else "
              f"aggregator-only. 'RE Gold-signal' = Residency Explorer interview rate WITH a Gold signal vs none "
              f"(shows how decisive a signal is). '◆' marks a resident/faculty from the applicant's own school ({same_school})."),
        ("S", f"Signals — {ng} Gold + {ns} Silver" + (f" (specialty allots {gold_n} Gold + {silver_n} Silver)" if gold_n else "")),
        ("T", f"GOLD: {golds or '(none set)'}."),
        ("T", f"SILVER: {silvers or '(none set)'}."),
        ("T", "Signals go to reach/target programs where they move the needle; apply to safeties without signaling."),
    ]
    if cfg.get("notes"):
        lines += [("S", "Strategy notes"), ("T", cfg["notes"])]
    lines += [
        ("W", "Verify before certifying"),
        ("T", "Aggregators (including Residency Explorer) can be wrong on visa and composition — the program "
              "website governs. Re-confirm the shortlist in Residency Explorer and on program sites, and check "
              "each program's years-since-graduation policy, before finalizing the ERAS list."),
    ]
    rr = 1
    for tag, txt in lines:
        c = ws0.cell(rr, 2, txt)
        if tag == "H": c.font = Font(bold=True, size=14, color=NAVY)
        elif tag == "S":
            c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = hf; ws0.row_dimensions[rr].height = 20
        elif tag == "W":
            c.font = Font(bold=True, size=11, color="9C1F1F"); c.fill = PatternFill("solid", fgColor="F5D9D0")
        else:
            c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
            ws0.row_dimensions[rr].height = 14 * (1 + len(txt) // 112)
        rr += 1

    ws = wb.create_sheet("Programs + Signals")
    cols = [("#", 4), ("Signal", 8), ("Division", 16), ("St", 5), ("Program", 40), ("Tier", 8),
            ("Type", 20), ("Non-US IMG", 20), (f"{affinity} residents", 34), ("Visa", 12),
            ("Fellowship / career", 24), ("RE Gold-signal", 16), ("Live ✓", 8), ("URL", 34)]
    ws.append([c[0] for c in cols])
    for i, o in enumerate(progs, 1):
        nm = str(g(o, "name")) + ("  ◆" if o.get("same_school") else "")
        ws.append([i, g(o, "signal"), g(o, "division"), g(o, "state"), nm, g(o, "tier", "Target"),
                   g(o, "type", g(o, "university")), g(o, "nonus_img"), g(o, "affinity"), g(o, "visa"),
                   g(o, "fellowship"), g(o, "re_gold"), "live ✓" if o.get("verified") else "", g(o, "url")])
    for c in range(1, len(cols) + 1):
        x = ws.cell(1, c); x.fill = hf; x.font = hfont
        x.alignment = Alignment(wrap_text=True, vertical="center"); x.border = bd
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30
    for i, (nm, wd) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    for rr in range(2, ws.max_row + 1):
        for cc in range(1, len(cols) + 1):
            x = ws.cell(rr, cc); x.alignment = wrap; x.border = bd; x.font = Font(size=8)
        sg = ws.cell(rr, 2).value
        if sg == "Gold": ws.cell(rr, 2).fill = GOLDF; ws.cell(rr, 2).font = Font(bold=True, size=8, color="7A5C00")
        elif sg == "Silver": ws.cell(rr, 2).fill = SILF; ws.cell(rr, 2).font = Font(bold=True, size=8, color="495260")
        tv = ws.cell(rr, 6).value
        if tv in tfill: ws.cell(rr, 6).fill = tfill[tv]
        if "H-1B" in str(ws.cell(rr, 10).value): ws.cell(rr, 10).fill = BLUE
        if str(ws.cell(rr, 13).value).startswith("live"): ws.cell(rr, 13).fill = LIVE
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    wb.save(a.out)
    print(f"Saved {a.out}  ({len(progs)} programs, {ng} Gold + {ns} Silver, {nlive} live-verified)")

if __name__ == "__main__":
    main()
