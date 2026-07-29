#!/usr/bin/env python3
"""Scaffold the per-project setup for an outcomes extraction batch.

Reads the pre-seeded template files and a folder of reference publications, then writes SKELETON
config files into the work dir so you don't hand-build the plumbing each time:

  - jobs.json            one entry per UNIQUE paper: matched source files + the needed
                         (table, family, comparison, subgroup, treatment, control) tuples
                         auto-collected from the pre-seeded rows. You fill in `design` (anchor
                         numbers from the abstract) and the default treatment/control arm guesses.
  - assemble_config.json results path + files map (with the detected family per file) + a
                         study_info skeleton (one row per paper; pmid pre-filled when the
                         template's Publication ID is already a PMID).
  - add_rows_config.json files map + an empty new_rows map (add extra comparison rows here; for
                         subgroup sheets, extra (comparison x subgroup) rows).

Usage:
  python scaffold.py --sources <references_dir> --out-dir _work \
      PWMA=/abs/pwma_template.xlsx NMA=/abs/nma_template.xlsx SUB=/abs/pwma_subgroup_template.xlsx

The table KEY is yours to choose (OS/DFS/RFS/PWMA/NMA/SUB…); the FAMILY is detected from the file's
header, never from the key or the filename. Each family has its own field-ID set, so nothing is
resolved by column letter.

Everything it writes is a STARTING POINT — review jobs.json (especially the file matches it flags as
ambiguous), fill the design notes and arm defaults, then run the extract workflow. The matching is a
convenience, not an oracle: confirm each file->paper mapping before extracting.
"""
import argparse, json, os, re, glob, sys
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import families as F


def col_text(ws, r, c):
    return "" if c is None else F.clean(ws.cell(row=r, column=c).value)


def first_author_tokens(author):
    """Name tokens (len>=3) of the FIRST author only, for matching against filenames.
    Handles ';'-separated ("De Placido, S.; Gallo, C.") and ','-separated ("Olivia Pagani, Barbara ...")
    author lists, and multi-token surnames ("De Placido", "Del Mastro") by keeping every token."""
    if not author: return []
    chunk = author.split(";")[0] if ";" in author else author.split(",")[0]
    return [t for t in re.findall(r"[A-Za-z]{3,}", chunk)]


def year_of(date_str):
    m = re.findall(r"(19|20)\d{2}", date_str or "")
    return m and re.search(r"((?:19|20)\d{2})", date_str).group(1) or ""


def is_pmid(pub_id):
    return bool(pub_id) and pub_id.isdigit()


def match_files(tokens, year, pub_id, sources):
    """Return (matched_files, note). Match source files by year + any first-author name token;
    fall back to token-only when the filename year differs from the publication year."""
    toks = [t.lower() for t in tokens]
    def tok_hit(base): return any(t in base for t in toks)
    cands = [f for f in sources if year and year in os.path.basename(f).lower() and tok_hit(os.path.basename(f).lower())]
    note = ""
    if not cands:                                   # filename year may differ from publish year
        cands = [f for f in sources if tok_hit(os.path.basename(f).lower())]
        note = "MATCHED BY AUTHOR ONLY (year differs) — verify" if cands else "NO FILE MATCH — fill manually"
    # group: keep all candidates (main + _suppl/_Sup/_sup1 etc.) sorted main-first
    cands = sorted(set(cands), key=lambda p: (len(os.path.basename(p)), os.path.basename(p)))
    # warn if only supplement-like files matched (main text likely named differently, e.g. by submission year)
    def is_suppl(b): return any(k in b.lower() for k in ("suppl", "_sup", "sup_", "appendix", "figure", "_table"))
    if cands and not note and all(is_suppl(os.path.basename(f)) for f in cands):
        note = "ONLY SUPPLEMENT FILES MATCHED — main text likely missing, verify"
    return cands, note


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sources", required=True, help="folder of reference PDFs/supplements")
    ap.add_argument("--out-dir", default="_work")
    ap.add_argument("tables", nargs="+", help="TABLEKEY=/abs/path.xlsx (key is yours; family is detected)")
    args = ap.parse_args()
    os.makedirs(args.out_dir, exist_ok=True)
    files = {}
    for spec in args.tables:
        k, p = spec.split("=", 1); files[k] = os.path.abspath(p)
    sources = [f for f in glob.glob(os.path.join(args.sources, "*")) if os.path.isfile(f)]

    # collect per-paper metadata + needed (table, comparison, subgroup) tuples
    papers = {}   # pid -> {author,year,pub_id, needed:[...]}
    fam_of, levels = {}, {}     # table -> family ; table -> {pid -> [subgroup levels]}
    for table, path in files.items():
        ws = F.sheet_of(openpyxl.load_workbook(path, data_only=True))
        lay = F.Layout(ws)
        fam_of[table] = lay.family
        if lay.missing_roles():
            print(f"WARNING [{table}]: header did not resolve roles {sorted(lay.missing_roles())}")
        for r in range(4, ws.max_row + 1):
            pid = ws.cell(row=r, column=lay.col_paper_id).value
            if pid is None: continue
            pid = int(pid)
            p = papers.setdefault(pid, {"author": col_text(ws, r, 2), "year": year_of(col_text(ws, r, 4)),
                                        "pub_id": col_text(ws, r, 5), "needed": []})
            comp = col_text(ws, r, lay.col_comparison) or "Primary"
            sub = col_text(ws, r, lay.col_subgroup) if lay.col_subgroup else ""
            p["needed"].append((table, lay.family, comp, sub))
            if sub:
                levels.setdefault(table, {}).setdefault(pid, [])
                if sub not in levels[table][pid]:
                    levels[table][pid].append(sub)

    jobs, study_info, flagged = [], {}, []
    for pid, p in sorted(papers.items()):
        tokens = first_author_tokens(p["author"])
        label = tokens[0] if tokens else "?"
        matched, note = match_files(tokens, p["year"], p["pub_id"], sources)
        if note: flagged.append(f"  {pid} ({label} {p['year']}): {note}")
        pmid = p["pub_id"] if is_pmid(p["pub_id"]) else ""
        needed = []
        for (t, fam, comp, sub) in p["needed"]:
            item = {"table": t, "family": fam, "comparison": comp,
                    "treatment": "FILL: experimental arm (T1 in nma)",
                    "control": "FILL: standard/comparator arm (T2 in nma)", "note": ""}
            if fam == "pwma_subgroup":
                item["subgroup"] = sub
            needed.append(item)
        jobs.append({
            "paper_id": pid, "trial": "", "pmid": pmid, "nct": "",
            "files": matched,
            "design": "FILL: trial design + anchor HRs/rates/events from the abstract",
            "_file_match_note": note or "matched by year+surname",
            "needed": needed,
        })
        study_info[str(pid)] = {"trial_name": "", "nct": "", "pmid": pmid, "arms": ""}

    out = lambda n: os.path.join(args.out_dir, n)
    files_cfg = {k: {"path": v, "family": fam_of[k]} for k, v in files.items()}
    json.dump(jobs, open(out("jobs.json"), "w"), indent=1)
    json.dump({"results": out("extraction_results.json"), "today": "YYYY-MM-DD",
               "files": files_cfg, "study_info": study_info}, open(out("assemble_config.json"), "w"), indent=1)
    json.dump({"files": files_cfg, "new_rows": {}}, open(out("add_rows_config.json"), "w"), indent=1)

    print(f"Scaffolded {len(jobs)} papers, {sum(len(j['needed']) for j in jobs)} required rows.")
    print(f"  wrote {out('jobs.json')}, {out('assemble_config.json')}, {out('add_rows_config.json')}")
    for k in files:
        n = sum(1 for j in jobs for x in j["needed"] if x["table"] == k)
        print(f"  {k}: family={fam_of[k]}, {n} rows")
        if k in levels:
            counts = sorted({len(v) for v in levels[k].values()})
            print(f"     subgroup levels per study: {counts} (VARIES — do not hardcode a number)")
            for pid, lv in levels[k].items():
                print(f"       {pid}: {len(lv)} -> {lv}")
    if flagged:
        print("FILE-MATCH FLAGS (review before extracting):")
        print("\n".join(flagged))
    else:
        print("All papers matched a source file set (still verify the matches).")


if __name__ == "__main__":
    main()
