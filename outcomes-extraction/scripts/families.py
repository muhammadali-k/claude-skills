#!/usr/bin/env python3
"""Template-family detection and header parsing, shared by every script in this skill.

There are THREE outcome-table families. They look almost identical and they are NOT interchangeable:

  pwma           A-Y (25 cols). Comparison label in F. Data G-Y (19 fields).
                 Et/Nt/Ec/Nc  ->  Et = EVENTS in treatment, Nt = N in treatment,
                                  Ec = EVENTS in control,   Nc = N in control.
                 (The legacy OS/DFS/RFS *_to_extract.xlsx sheets are this family.)

  nma            A-X (24 cols). Comparison label in F. Data G-X (18 fields; no "(0 selected)").
                 Ec T1/Et T1/Ec T2/Et T2  ->  Ec = EVENT COUNT, Et = EVENT TOTAL (participants).
                 T1 = the treatment arm, T2 = the active comparator.

  pwma_subgroup  A-AA (27 cols). Subgroup label in F, "Extraction Possible" in G,
                 comparison label in H. Data I-AA (19 fields). Same Et/Nt/Ec/Nc meaning as pwma.

***  Et MEANS DIFFERENT THINGS IN pwma AND nma.  ***
In pwma/pwma_subgroup, `Et` is an EVENT COUNT. In nma, `Et` is a PARTICIPANT TOTAL (the denominator)
and `Ec` is the event count. Carrying one convention into the other silently swaps numerator and
denominator and the sheet still looks plausible. The LABEL->ROLE tables below are the single place
that mapping is defined; every script resolves counts through them and never by column letter.

The internal role vocabulary is deliberately unambiguous — `events_t`, `n_t`, `events_c`, `n_c` —
precisely so that no code anywhere has to think about what "Et" means.
"""
import re

# Missing-value tokens seen across the real templates: the PWMA seed uses NA/NR/NM,
# the subgroup seed uses N/A. Treat them all as "not reported" when reading.
MISSING_TOKENS = {"", "na", "n/a", "nr", "nm", "n.r.", "none", "not reported", "not stated"}

# role -> header label, per family. Keys prefixed "GROUP|" are disambiguated by the row-2
# group banner because the row-3 label alone is ambiguous (nma prints "T1"/"T2" twice:
# once under "Regimen" for the arm name, once under "Survival" for the landmark rate).
LABELS = {
    "pwma": {
        "trial name": "trial_name", "nct": "nct", "pmid": "pmid", "o/f": "of",
        "(0 selected)": "zero", "arms": "arms",
        "te": "te", "lower ci": "lower_ci", "upper ci": "upper_ci",
        "treatment": "treatment", "control": "control",
        "survival in treatment": "surv_t", "survival in control": "surv_c",
        "et": "events_t", "nt": "n_t", "ec": "events_c", "nc": "n_c",
        "median survival in treatment": "med_t", "median survival in control": "med_c",
    },
    "nma": {
        "trial name": "trial_name", "nct": "nct", "pmid": "pmid", "o/f": "of", "arms": "arms",
        "measure": "te", "lower ci": "lower_ci", "upper ci": "upper_ci",
        "regimen|t1": "treatment", "regimen|t2": "control",
        "survival|t1": "surv_t", "survival|t2": "surv_c",
        # THE TRAP, spelled out: in nma, Ec is the numerator and Et is the denominator.
        "ec t1": "events_t", "et t1": "n_t", "ec t2": "events_c", "et t2": "n_c",
        "median survival in treatment": "med_t", "median survival in control": "med_c",
    },
}
LABELS["pwma_subgroup"] = dict(LABELS["pwma"])

# Fields the extraction fills (as opposed to per-study metadata assemble copies in).
DATA_ROLES = ("te", "lower_ci", "upper_ci", "treatment", "control",
              "surv_t", "surv_c", "events_t", "n_t", "events_c", "n_c", "med_t", "med_c")
META_ROLES = {
    "pwma": ("trial_name", "nct", "pmid", "of", "zero", "arms"),
    "pwma_subgroup": ("trial_name", "nct", "pmid", "of", "zero", "arms"),
    "nma": ("trial_name", "nct", "pmid", "of", "arms"),
}

# Which sheet-column letters hold the counts, per family — for error messages only.
COUNT_COLS_HELP = {
    "pwma": "T=Et(events,treatment) U=Nt(N,treatment) V=Ec(events,control) W=Nc(N,control)",
    "pwma_subgroup": "V=Et(events,treatment) W=Nt(N,treatment) X=Ec(events,control) Y=Nc(N,control)",
    "nma": "S=Ec T1(events,T1) T=Et T1(N,T1) U=Ec T2(events,T2) V=Et T2(N,T2)",
}


def sheet_of(wb):
    """The outcome sheet. All three templates name it 'Outcome Table'; fall back to active."""
    for name in wb.sheetnames:
        if name.strip().lower() == "outcome table":
            return wb[name]
    return wb.active


def clean(v):
    return "" if v is None else " ".join(str(v).split())


def label_of(v):
    """Header label with the '(ID: NNNN)' tag stripped, lowercased."""
    return re.split(r"\(ID", clean(v))[0].strip().lower()


def id_of(v):
    m = re.search(r"\(ID:\s*(\d+)\)", clean(v))
    return m.group(1) if m else None


def banner(ws, row):
    """{col: label} for a header row, expanding merged ranges so a merged group banner
    (e.g. 'Regimen' over L2:M2) is visible on every column it covers."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = clean(ws.cell(row=row, column=c).value)
        if v:
            out[c] = v
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row:
            v = clean(ws.cell(row=mr.min_row, column=mr.min_col).value)
            if v:
                for c in range(mr.min_col, mr.max_col + 1):
                    out.setdefault(c, v)
    return out


class Layout:
    """Parsed header of one outcome sheet: family, role->column, role->field-id, key columns."""

    def __init__(self, ws):
        self.ws = ws
        r1, r2 = banner(ws, 1), banner(ws, 2)

        # --- key columns, located by their row-1 label (never hardcoded letters) ---
        idx = {label_of(v): c for c, v in r1.items()}
        self.col_paper_id = idx.get("paper id", 1)
        self.col_subgroup = idx.get("subgroup")
        self.col_extraction_possible = idx.get("extraction possible")
        self.col_comparison = idx.get("treatment arm")
        # data block = everything the row-1 'Root Node' banner covers
        root = [c for c, v in r1.items() if label_of(v) == "root node"]
        self.data_first_col = min(root) if root else (self.col_comparison or 6) + 1
        self.data_last_col = max(root) if root else ws.max_column

        # --- collect (group, label, id) per data column ---
        cols = []
        for c in range(self.data_first_col, self.data_last_col + 1):
            lab2, lab3 = ws.cell(row=2, column=c).value, ws.cell(row=3, column=c).value
            lab = label_of(lab3) or label_of(lab2)
            fid = id_of(lab3) or id_of(lab2)
            grp = label_of(r2.get(c, "")) if label_of(lab3) else ""
            cols.append((c, grp, lab, fid))

        labels = {lab for _, _, lab, _ in cols}
        if self.col_extraction_possible:
            self.family = "pwma_subgroup"
        elif "ec t1" in labels or "measure" in labels:
            self.family = "nma"
        else:
            self.family = "pwma"

        table = LABELS[self.family]
        self.roles, self.ids, self.labels, self.unmapped = {}, {}, {}, []
        for c, grp, lab, fid in cols:
            role = table.get(f"{grp}|{lab}") or table.get(lab)
            if role is None:
                self.unmapped.append((c, grp, lab))
                continue
            if role not in self.roles:      # first occurrence wins
                self.roles[role] = c
                self.ids[role] = fid
                self.labels[role] = f"{grp}|{lab}" if grp and grp != lab else lab

    # --- helpers -----------------------------------------------------------
    def expected_cols(self):
        return self.data_last_col

    def required_roles(self):
        return set(META_ROLES[self.family]) | set(DATA_ROLES)

    def missing_roles(self):
        return self.required_roles() - set(self.roles)

    def col(self, role):
        return self.roles.get(role)

    def key(self, r):
        """Row identity: (paper_id, comparison label, subgroup label). Subgroup is '' for
        the main sheets, so one keying scheme covers all three families."""
        pid = self.ws.cell(row=r, column=self.col_paper_id).value
        comp = clean(self.ws.cell(row=r, column=self.col_comparison).value) if self.col_comparison else ""
        sub = clean(self.ws.cell(row=r, column=self.col_subgroup).value) if self.col_subgroup else ""
        return pid, comp, sub

    def meta_cols(self):
        """The pre-seeded left-hand columns (A .. last key column) copied when adding rows."""
        last = max(x for x in (self.col_paper_id, self.col_comparison, self.col_subgroup,
                               self.col_extraction_possible) if x)
        return list(range(1, last + 1))

    def describe(self):
        from openpyxl.utils import get_column_letter as L
        bits = [f"family={self.family}", f"cols=A-{L(self.data_last_col)} ({self.data_last_col})",
                f"paper_id={L(self.col_paper_id)}"]
        if self.col_subgroup:
            bits.append(f"subgroup={L(self.col_subgroup)}")
        if self.col_extraction_possible:
            bits.append(f"extraction_possible={L(self.col_extraction_possible)}")
        if self.col_comparison:
            bits.append(f"comparison={L(self.col_comparison)}")
        bits.append(f"data={L(self.data_first_col)}-{L(self.data_last_col)}")
        return "  ".join(bits)


def files_from_cfg(cfg):
    """`files` in a config may be {"KEY": "/path.xlsx"} (legacy) or
    {"KEY": {"path": "/path.xlsx", "family": "nma"}} (what scaffold.py now writes).
    Returns {key: path}. The declared family is informational only — every script re-detects the
    family from the file's own header, because the header is the truth."""
    out = {}
    for k, v in (cfg.get("files") or {}).items():
        out[k] = v["path"] if isinstance(v, dict) else v
    return out


def parse(path_or_ws):
    """Layout for a path or an already-open worksheet."""
    if isinstance(path_or_ws, str):
        import openpyxl
        return Layout(sheet_of(openpyxl.load_workbook(path_or_ws, data_only=True)))
    return Layout(path_or_ws)


def is_missing(v):
    return clean(v).lower() in MISSING_TOKENS


def num(v):
    """Float, or None when the cell is missing/non-numeric."""
    s = clean(v).replace(",", "").replace("%", "")
    if not s or is_missing(v):
        return None
    try:
        return float(s)
    except ValueError:
        return None
