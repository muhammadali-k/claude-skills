#!/usr/bin/env python3
"""QC filled outcome tables — all three families (pwma, nma, pwma_subgroup).

Usage:
  python qc.py <filled.xlsx> [<filled.xlsx> ...] [--vocabulary <project>_node_vocabulary.json]

Checks per file:
  - the header resolves to a known family and every expected role is present
  - the sheet has the expected number of columns for that family (pwma 25 / nma 24 / subgroup 27)
  - no empty data cell (every field filled, "NA" where unreported)
  - NA fill-rate per column
  - effect sanity: lower <= TE <= upper (TE = "TE" in pwma, "Measure" in nma)
  - a CI with no point estimate (almost always a survival-RATE CI mis-placed into the effect CI cells)
  - *** EVENTS > DENOMINATOR *** — the Ec/Et guard, see below
  - subgroup sheets: "Extraction Possible" answered, no duplicate (study, comparison, subgroup) rows,
    and no row that says extraction is impossible while still carrying an effect estimate
  - *** NODE LABELS *** — with --vocabulary, every treatment- and control-arm label is resolved against
    the project's controlled vocabulary and the RUN FAILS (exit 1) on any rejection, see below

*** NODE LABELS (--vocabulary) ***
netmeta joins arms by STRING EQUALITY of their labels. "NIVO + IPI" and "NIVO+IPI" are two nodes, not
one: the network fragments, or worse stays connected with an edge missing, and nothing in this sheet
looks wrong — the failure only surfaces as a wrong league table. So every treatment/control label is
resolved through `nodes.py` against the project's vocabulary file. Rejections print the validator's own
diagnosis (which carries a "did you mean" suggestion) and fail the run. An ALIAS hit — a legacy or
paper-prose spelling the project has explicitly mapped — is accepted but REPORTED, never silent, so a
sheet full of legacy labels is visible as such. Without --vocabulary the labels are not checked at all
and the script says so.

*** THE Ec/Et GUARD (the reason this check exists) ***
`Et` means opposite things in the two pairwise/network layouts:
    pwma / pwma_subgroup : Et = EVENTS in treatment, Nt = N in treatment,
                           Ec = EVENTS in control,   Nc = N in control
    nma                  : Ec T1 = EVENT COUNT in T1, Et T1 = EVENT TOTAL (participants) in T1
                           Ec T2 / Et T2 likewise for the comparator
So an extractor who carries the PWMA convention into an NMA sheet writes the participant total where
the event count belongs and vice versa. The output still looks plausible and NOTHING downstream
catches it. The one arithmetic fact that always breaks is events > denominator, so this script flags
any row where the event count exceeds its own arm's N, per family, and says which two columns to swap.
"""
import argparse, sys, os, collections, openpyxl
from openpyxl.utils import get_column_letter as L

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import families as F
import nodes as N

NO_TOKENS = {"no", "n", "false", "0", "not possible", "not extractable"}

NO_VOCAB_WARNING = """  WARNING: NODE LABELS ARE UNVALIDATED — no --vocabulary was given.
      netmeta joins arms by string equality, so "NIVO + IPI", "Nivo+Ipi" and "NIVO+IPI" are three
      separate nodes. The sheet still looks correct in every cell; the damage shows up only in the
      league table, after the analysis has been run and read. Nothing else in this script can catch it.
      Re-run with --vocabulary <project>_node_vocabulary.json."""


def qc(path, vocab=None):
    """Returns the number of REJECTED node labels (0 when --vocabulary was not supplied)."""
    wb = openpyxl.load_workbook(path, data_only=True); ws = F.sheet_of(wb)
    lay = F.Layout(ws)
    print(f"\n=== {path}")
    print(f"    {lay.describe()}")
    print(f"    counts: {F.COUNT_COLS_HELP[lay.family]}")
    if lay.missing_roles():
        print(f"  ERROR: header did not resolve roles {sorted(lay.missing_roles())}")
    if ws.max_column != lay.expected_cols():
        print(f"  WARNING: expected {lay.expected_cols()} columns for family {lay.family}, "
              f"got {ws.max_column}")

    empties, na_count, hr_issues, ci_no_te = [], {}, [], []
    count_issues, ep_missing, ep_contradiction, dup = [], [], [], []
    node_rows = []
    seen = collections.Counter()
    nrows = 0

    for r in range(4, ws.max_row + 1):
        pid, comp, sub = lay.key(r)
        if pid is None: continue
        nrows += 1
        tag = f"row{r} pid={pid} comp={comp!r}" + (f" sub={sub!r}" if sub else "")
        seen[(pid, F.clean(comp).lower(), F.clean(sub).lower())] += 1

        for c in range(lay.data_first_col, lay.data_last_col + 1):
            v = ws.cell(row=r, column=c).value
            col = L(c)
            if v is None or str(v).strip() == "":
                empties.append((r, col, pid))
            elif F.is_missing(v):
                na_count[col] = na_count.get(col, 0) + 1

        cell = lambda role: ws.cell(row=r, column=lay.col(role)).value if lay.col(role) else None

        # --- node labels, collected here and resolved after the loop (only with --vocabulary) ---
        for role in ("treatment", "control"):
            if lay.col(role):
                node_rows.append({"record": tag, "pid": pid, "field": role, "value": cell(role)})

        te, lo, up = F.num(cell("te")), F.num(cell("lower_ci")), F.num(cell("upper_ci"))
        if None not in (te, lo, up) and not (lo <= te <= up):
            hr_issues.append((tag, lo, te, up))
        # a reported CI always implies a point estimate; a CI with no TE is a mis-filed RATE CI
        if te is None and (lo is not None or up is not None):
            ci_no_te.append((tag, lo, up))

        # --- events vs denominator, per arm ---
        for ev_role, n_role, arm in (("events_t", "n_t", "treatment/T1"), ("events_c", "n_c", "control/T2")):
            ev, n = F.num(cell(ev_role)), F.num(cell(n_role))
            if ev is None or n is None: continue
            if ev > n:
                ecol, ncol = L(lay.col(ev_role)), L(lay.col(n_role))
                hint = (f"swap {ecol}<->{ncol} — looks like the pwma Et/Ec convention was carried into "
                        f"an nma sheet" if lay.family == "nma" else
                        f"swap {ecol}<->{ncol} — looks like the nma Ec/Et convention was carried into "
                        f"a pwma sheet")
                count_issues.append((tag, arm, f"{lay.labels.get(ev_role)}={ev:g} > "
                                                f"{lay.labels.get(n_role)}={n:g}", hint))

        # --- subgroup-sheet specific ---
        if lay.col_extraction_possible:
            ep = F.clean(ws.cell(row=r, column=lay.col_extraction_possible).value)
            if not ep or F.is_missing(ep):
                ep_missing.append(tag)
            elif ep.lower() in NO_TOKENS and te is not None:
                ep_contradiction.append((tag, ep, te))
            if not F.clean(sub):
                dup.append((tag, "blank subgroup label"))

    for key, n in seen.items():
        if n > 1:
            dup.append((f"pid={key[0]} comp={key[1]!r} sub={key[2]!r}", f"{n} duplicate rows"))

    print(f"  data rows: {nrows}")
    print(f"  EMPTY cells (should be 0): {len(empties)}", empties[:10] if empties else "")
    print(f"  missing/NA per column: " + ", ".join(f"{k}:{v}" for k, v in sorted(na_count.items())))
    print(f"  effect sanity violations (need lower<=TE<=upper): {len(hr_issues)}", hr_issues or "")
    print(f"  CI-without-TE (likely mis-placed rate CI): {len(ci_no_te)}", ci_no_te or "")
    print(f"  *** EVENTS > DENOMINATOR: {len(count_issues)} ***")
    for tag, arm, detail, hint in count_issues:
        print(f"      {tag}  [{arm}]  {detail}\n         -> {hint}")
    if lay.col_extraction_possible:
        print(f"  'Extraction Possible' unanswered: {len(ep_missing)}", ep_missing[:10] or "")
        print(f"  'Extraction Possible'=no but an effect estimate is present: {len(ep_contradiction)}",
              ep_contradiction or "")
    print(f"  duplicate / unlabelled row keys: {len(dup)}", dup or "")

    return check_nodes(vocab, node_rows)


def check_nodes(vocab, node_rows):
    """Resolve every treatment/control label against the vocabulary. Returns the rejection count.

    Rejections are printed with the validator's own diagnosis, VERBATIM — it already names the rule
    that was broken and suggests the canonical string, and rewording it here would put a second,
    drifting copy of the rules in a second file."""
    if vocab is None:
        print(NO_VOCAB_WARNING)
        return 0

    ok, problems, resolved = N.check_rows(vocab, node_rows)
    rejected = [p for p in problems if not p["problem"].startswith("ACCEPTED")]
    aliased = [p for p in problems if p["problem"].startswith("ACCEPTED")]

    print(f"  *** NODE LABELS: {len(node_rows)} checked, {len(rejected)} REJECTED, "
          f"{len(aliased)} accepted-but-non-canonical ***")
    for p in rejected:
        print(f"      REJECTED  {p['record']} [{p['field']}]  {p['problem']}")
    for p in aliased:
        # An alias hit is a pass, but never a silent one: it means the sheet carries a legacy or
        # paper-prose spelling that only survives because the project mapped it.
        print(f"      {p['record']} [{p['field']}]  {p['problem']}")

    # The vocabulary's `arms` map pins which labels belong to which record. A label can be perfectly
    # well-formed and still be attached to the wrong trial; that is reported, not failed, because the
    # map is a project bookkeeping aid rather than part of the label rules.
    misplaced = []
    for row in node_rows:
        label = resolved.get((row["record"], row["field"]))
        expected = vocab.arms.get(str(row["pid"]))
        if label is None or not expected:
            continue
        want = expected.get("treatment", []) if row["field"] == "treatment" else [expected.get("control")]
        if label not in want:
            misplaced.append(f"{row['record']} [{row['field']}] = {label} but `arms` for this record "
                             f"says {want} ({expected.get('trial', '?')})")
    print(f"  node labels not matching the vocabulary's `arms` map (not a failure): {len(misplaced)}")
    for m in misplaced:
        print(f"      {m}")
    return len(rejected)


def main():
    ap = argparse.ArgumentParser(description="QC filled outcome tables (pwma / nma / pwma_subgroup).")
    ap.add_argument("files", nargs="+", help="filled template .xlsx file(s)")
    ap.add_argument("--vocabulary", help="project *_node_vocabulary.json; without it node labels are "
                                         "NOT validated and a wrong label is invisible until the "
                                         "league table is wrong")
    a = ap.parse_args()

    vocab = N.Vocabulary.load(a.vocabulary) if a.vocabulary else None
    if vocab is not None:
        print(f"Node vocabulary: {a.vocabulary}  (comparator={vocab.comparator}, "
              f"{len(vocab.agents)} agents, {len(vocab.combinations)} combinations, "
              f"{len(vocab.variants)} variants, {len(vocab.aliases)} aliases)")

    rejected = sum(qc(p, vocab) for p in a.files)
    print("\nQC done.")
    if rejected:
        sys.exit(f"FAILED: {rejected} node label(s) rejected. Every one of them would have become a "
                 f"separate node in the network. Fix the labels (or add the alias/variant/combination "
                 f"to the vocabulary) and re-run.")


if __name__ == "__main__":
    main()
