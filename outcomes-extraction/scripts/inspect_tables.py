#!/usr/bin/env python3
"""Dump the structure of any outcome template (pwma / nma / pwma_subgroup) so you learn the layout
from the FILE rather than from memory.

Usage:
  python inspect_tables.py <file.xlsx> [<file.xlsx> ...]

For each sheet it prints: the detected TEMPLATE FAMILY, the key columns (paper id / subgroup /
extraction-possible / comparison), the role -> column -> field-ID map (with the group banner used to
disambiguate the nma "T1"/"T2" duplicates), the counts semantics for that family, any unmapped
header, and the data rows. For subgroup sheets it also prints the subgroup levels per study — the
count VARIES by study, so never assume a fixed number.

Read this before extracting. The field IDs differ per template (nma 21972-21993, pwma 3069-3089,
subgroup 9432-9452, legacy OS/DFS/RFS 3223-3287 …), which is exactly why nothing maps by column
letter.
"""
import sys, os, collections, openpyxl
from openpyxl.utils import get_column_letter as L

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import families as F


def dump(path):
    print("=" * 96)
    print("FILE:", path)
    print("=" * 96)
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        lay = F.Layout(ws)
        print(f"\n--- SHEET: {ws.title}  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
        print("  " + lay.describe())
        print("  COUNTS SEMANTICS: " + F.COUNT_COLS_HELP[lay.family])
        if lay.family == "nma":
            print("  !! nma: Ec = EVENT COUNT, Et = EVENT TOTAL (participants). T1 = treatment arm,")
            print("     T2 = active comparator. This is the OPPOSITE of the pwma Et/Ec meaning.")
        else:
            print("  !! pwma: Et = EVENTS in treatment, Ec = EVENTS in control; N lives in Nt/Nc.")
        if ws.max_column != lay.expected_cols():
            print(f"  WARNING: data block ends at {L(lay.data_last_col)} but sheet has "
                  f"{ws.max_column} columns")

        print("  role -> column (field ID) [header label]:")
        for role, c in sorted(lay.roles.items(), key=lambda kv: kv[1]):
            print(f"    {role:<12} {L(c):>2}  (ID: {lay.ids.get(role)})  [{lay.labels.get(role)}]")
        if lay.missing_roles():
            print("  MISSING ROLES (header did not resolve):", sorted(lay.missing_roles()))
        if lay.unmapped:
            print("  UNMAPPED HEADERS:", [(L(c), g, l) for c, g, l in lay.unmapped])

        # data rows
        print("  --- data rows ---")
        per_study = collections.defaultdict(list)
        for r in range(4, ws.max_row + 1):
            pid, comp, sub = lay.key(r)
            if pid is None:
                continue
            if sub:
                per_study[pid].append((comp, sub))
            cells = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    cells.append(f"{L(c)}{r}={v!r}")
            if cells:
                print("   " + " | ".join(cells))

        if lay.col_subgroup:
            print("  --- subgroup levels per study (VARIES — do not hardcode) ---")
            for pid, items in per_study.items():
                by_comp = collections.defaultdict(list)
                for comp, sub in items:
                    by_comp[comp].append(sub)
                for comp, subs in by_comp.items():
                    print(f"    {pid} | comparison={comp!r}: {len(subs)} level(s) -> {subs}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: python inspect_tables.py <file.xlsx> [<file.xlsx> ...]")
    for p in sys.argv[1:]:
        dump(p)
