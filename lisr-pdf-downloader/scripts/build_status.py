#!/usr/bin/env python3
"""Build Download_status.xlsx from resolved.json + oa_results.json + browser_results.json + files.

Columns: Study ID, Status, File, Size(KB), Supplements, Identifier, ID type, Title, Journal, Year,
Match score, DOI, Action link (clickable LibKey / ClinicalTrials.gov), Notes. Plus a Summary sheet.

Usage:
  python build_status.py --out Downloaded_files --resolved resolved.json --xlsx Download_status.xlsx --library 964
"""
import argparse, json, os, re, urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils as U

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default='Downloaded_files')
    ap.add_argument('--resolved', default='resolved.json')
    ap.add_argument('--xlsx', default='Download_status.xlsx')
    ap.add_argument('--library', default='964')
    a = ap.parse_args()

    resolved = json.load(open(a.resolved, encoding='utf-8'))
    br = json.load(open('browser_results.json', encoding='utf-8')) if os.path.exists('browser_results.json') else {}
    files = {f[:-4]: os.path.getsize(os.path.join(a.out, f)) for f in os.listdir(a.out)
             if f.lower().endswith('.pdf') and '_suppl' not in f}
    supp = {}
    for f in os.listdir(a.out):
        if '_suppl' in f.lower():
            base = f.lower().split('_suppl')[0]
            supp.setdefault(base, []).append(f)
    ids_order = list(resolved.keys())

    def category(sid):
        rec = resolved.get(sid, {})
        if sid in files and files[sid] > 2500: return 'Downloaded'
        if rec.get('resolve_status') == 'trial_registration': return 'Trial registration (no journal article)'
        if not rec.get('doi'): return 'Full text unavailable - no DOI (likely conference abstract)'
        bs = br.get(sid, {}).get('status', ''); doi = rec.get('doi', '')
        if bs.startswith('CF') or doi.split('/')[0] in ('10.1016', '10.2139') or 'annonc' in doi:
            return 'Needs manual - publisher bot-wall (Cloudflare)'
        if bs.startswith('LIBKEY'): return 'Full text unavailable - no institutional access via LibKey'
        if bs.startswith('FAIL'): return 'Needs manual - publisher-specific download issue'
        return 'Not yet attempted'

    def action_link(sid):
        rec = resolved.get(sid, {})
        if sid in files and files[sid] > 2500: return ''
        if rec.get('nct'): return 'https://clinicaltrials.gov/study/' + rec['nct']
        if rec.get('doi'): return f'https://libkey.io/libraries/{a.library}/' + urllib.parse.quote(rec['doi'], safe='/')
        return ''

    wb = openpyxl.Workbook(); s = wb.active; s.title = 'Download_status'
    headers = ['Study ID', 'Status', 'File', 'Size (KB)', 'Supplements', 'Identifier', 'ID type',
               'Title', 'Journal', 'Year', 'Match score', 'DOI', 'Action link (click to fetch)', 'Notes']
    s.append(headers)
    hdr = PatternFill('solid', fgColor='1F4E78'); thin = Side(style='thin', color='D9D9D9')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in s[1]: c.font = Font(bold=True, color='FFFFFF'); c.fill = hdr; c.alignment = Alignment(vertical='center', wrap_text=True); c.border = bd
    green = PatternFill('solid', fgColor='C6EFCE'); red = PatternFill('solid', fgColor='FFC7CE')
    yellow = PatternFill('solid', fgColor='FFEB9C'); grey = PatternFill('solid', fgColor='E0E0E0')

    from collections import Counter
    cc = Counter()
    for sid in ids_order:
        rec = resolved.get(sid, {}); cat = category(sid); cc[cat] += 1
        fname = f'{sid}.pdf' if (sid in files and files[sid] > 2500) else ''
        size = round(files[sid] / 1024, 1) if fname else ''
        suppl = ', '.join(sorted(supp.get(sid, []))) if sid in supp else ''
        yr = ''; m = re.search(r'(19|20)\d{2}', str(rec.get('publish_date') or ''))
        if m: yr = m.group(0)
        score = rec.get('match_score') if rec.get('category') in ('pmid', 'embase_title') else ''
        notes = []
        if rec.get('resolve_status') == 'resolved_doi_check': notes.append('DOI match: verify')
        if rec.get('nct'): notes.append('NCT ' + rec['nct'])
        bs = br.get(sid, {}).get('status', '')
        if bs.startswith('FAIL') or bs.startswith('CF'): notes.append(bs)
        link = action_link(sid)
        s.append([sid, cat, fname, size, suppl, rec.get('paper_id') or '', rec.get('paper_id_type') or '',
                  rec.get('title') or '', rec.get('journal') or '', yr, score, rec.get('doi') or '', link, '; '.join(notes)])
        r = s[s.max_row]
        r[1].fill = green if cat == 'Downloaded' else (yellow if cat.startswith('Needs manual') else (grey if 'Trial registration' in cat else red))
        if link:
            lc = r[12]; lc.hyperlink = link; lc.font = Font(color='0563C1', underline='single')
        for c in r: c.border = bd; c.alignment = Alignment(vertical='top', wrap_text=True)
    widths = [10, 42, 13, 9, 26, 22, 9, 44, 28, 6, 7, 28, 40, 22]
    for i, w in enumerate(widths, 1): s.column_dimensions[U.get_column_letter(i)].width = w
    s.freeze_panes = 'A2'

    s2 = wb.create_sheet('Summary'); s2.append(['Outcome', 'Count'])
    for c in s2[1]: c.font = Font(bold=True, color='FFFFFF'); c.fill = hdr
    for k, v in sorted(cc.items(), key=lambda x: -x[1]): s2.append([k, v])
    s2.append(['TOTAL', sum(cc.values())])
    s2.column_dimensions['A'].width = 52; s2.column_dimensions['B'].width = 10
    wb.save(a.xlsx)
    print('=== SUMMARY ===')
    for k, v in sorted(cc.items(), key=lambda x: -x[1]): print(f'  {v:3}  {k}')
    print(f'  ---\n  {sum(cc.values())}  TOTAL  -> {a.xlsx}')

if __name__ == '__main__':
    main()
