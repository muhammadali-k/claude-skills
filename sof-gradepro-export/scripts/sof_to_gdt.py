#!/usr/bin/env python3
"""Convert per-outcome Summary-of-Findings Excel exports into one combined
GRADEpro-GDT-format JSON-LD file for MAGICapp's Evidence tab -> "Add PICO" ->
"Import PICO using a GDT Gradepro file" beta feature.

Usage:
    python3 sof_to_gdt.py <xlsx-file-or-folder> \
        --population "..." --intervention "..." --comparator "..." \
        [--title "..."] [--sof-title "..."] [--outcome-name "..."] \
        [--mid-small PCT] [--mid-moderate PCT] [--mid-large PCT] \
        [--favorable-direction reduction|increase] [--no-pls] \
        -o <output.json>

Input is either ONE outcome's .xlsx (single sheet "Summary of Findings", headers in
rows 1-2, one data record starting row 3, citations one-per-row overflowing into
extra rows below) or a FOLDER of many such files -- one outcome per file, combined
into a single PICO/question with one evidenceSummary entry per outcome, in the
order files are found (sorted by filename).

Output:
    <output.json>                                combined GDT JSON-LD, one PICO
    <output-stem>_post_import_checklist.md        one section per outcome: the exact
                                                   intervention-arm rate to expect after
                                                   clicking "Calculate estimates" in
                                                   MAGICapp, the recommended "Direction
                                                   of benefit" selection, and the
                                                   plain-language sentence to paste --
                                                   see LIMITATIONS below for why these
                                                   three are a required manual step

Non-obvious rules / documented limitations (see also the skill's SKILL.md):
  - Relative-effect prefix (HR/RR/OR/...) is parsed with a generic regex, not
    hardcoded to "HR" -- unrecognized prefixes still convert, using the literal
    prefix text as the relativeEffect @type, with a printed warning. The absolute-
    effect CI transform (see below) only runs for HR/RR/OR specifically; other
    prefixes still convert the point values but leave the CI transform null.
  - Risk-difference sign ("fewer"/"more") sets the absoluteEffect value's sign.
  - controlRisk.value = the control arm's "X per 1000" figure / 10 (GDT stores
    control risk per 100, confirmed against a real GDT export).
  - absoluteEffect.confidenceIntervalFrom/To ARE NOW COMPUTED (previously always
    null) using the standard GRADE/Cochrane relative-to-absolute transform, applied
    to the relative effect's own CI bounds against the control risk:
      HR:  interventionRisk = 1 - (1 - controlRisk) ** HR   (constant-hazard/
           exponential-survival transform)
      RR:  interventionRisk = controlRisk * RR
      OR:  interventionOdds = (controlRisk / (1 - controlRisk)) * OR;
           interventionRisk = interventionOdds / (1 + interventionOdds)
    This is not a guess -- it was reverse-engineered by live-testing MAGICapp's own
    "Calculate estimates" button (Absolute effect estimates -> Auto-calculated) against
    a real outcome and confirmed to reproduce its output to the reported precision (the
    live test used real project data, not reproduced here; illustrating with a fictional
    equivalent: HR 0.60, 95% CI 0.35-1.05, control 650/1000 -> this formula gives
    intervention ~467/1000, i.e. ~183 fewer per 1000, with a CI-bound difference range
    of roughly 343 fewer to 18 more per 1000).
  - IMPORTANT: MAGICapp's beta GDT importer does NOT run this calculation itself on
    import, even though it correctly imports the two inputs (controlRisk and
    relativeEffect) needed to run it. Confirmed live: after importing a JSON-LD file
    with a fully-populated absoluteEffect (including the CI bounds this script now
    computes), the outcome table's intervention-arm cell was still blank until
    "Calculate estimates" was clicked by hand in MAGICapp's UI. There is no field in
    GRADEpro's own JSON-LD export schema for this (confirmed absent from a real GDT
    sample export) -- "auto-calculate the absolute effect on import" is apparently a
    MAGICapp UI action, not something a GDT/JSON-LD file can trigger. So: this script
    still computes and writes the correct numbers into the JSON (harmless, and
    possibly read by a future, less-beta version of the importer), AND the companion
    checklist tells you/your collaborator the exact number to expect so clicking
    "Calculate estimates" for each outcome after import is a fast confirm-not-guess
    step, not a blind one.
  - patientGroup totalCount (N per arm) is always left null: not present anywhere
    in the source Excel.
  - studyDesign is always the generic {"RandomisedTrials","randomised trials"}
    default -- the source Excel does not state study design.
  - quality domain sub-ratings (riskOfBias/inconsistency/indirectness/imprecision)
    are always null: the source Excel gives only the single overall "Overall: X"
    rating, no domain breakdown.
  - controlRisk's @type label (Low/Moderate/HighControlRisk) is cosmetic/arbitrary
    in GDT's own schema -- this script always emits "LowControlRisk" and does not
    try to infer a meaningful band from the control rate.
  - Plain-language summaries are generated deterministically (GRADE Table 15.6.b /
    Cochrane Handbook Ch.15 conventions -- certainty-keyed verb choice + magnitude
    classification, including the CI-crosses-null special cases) and are (a) best-
    effort embedded into the output JSON-LD's explanation[] array, one item per
    outcome with a non-standard "forOutcome" cross-reference (there is no confirmed
    native per-outcome slot in this JSON-LD schema for this content, and MAGICapp's
    own help docs confirm the GDT importer does not import plain-language text even
    when present), and (b) ALWAYS also written to the companion checklist file, which
    is the reliable path -- paste those sentences into each outcome's "Plain language
    summary" field in MAGICapp by hand after import.
  - "Direction of benefit" (MAGICapp's own field: Intervention favourable /
    Comparator favourable / No important difference / High uncertainty) is likewise
    NOT settable via the GDT JSON-LD import -- confirmed absent from a real GDT
    sample export, and this is a MAGICapp-only concept GRADEpro's schema has no term
    for at all. This script classifies the recommended value deterministically (same
    crosses-null / magnitude logic as the plain-language generator) and writes it to
    the companion checklist for manual selection. The classification ASSUMES this
    skill's v1 scope (dichotomous/time-to-event outcomes: OS/DFS/PFS/RFS-style,
    "event" = an undesirable thing happening, e.g. death/progression/recurrence) --
    i.e. a reduction in the event rate is assumed favourable to the intervention.
    Override with --favorable-direction increase for the rare outcome where a higher
    rate is the desirable direction; there is no per-outcome override, only a
    per-run flag (documented limitation, matching --mid-small/--mid-moderate/--mid-large).
  - Magnitude bands (trivial/small/moderate/large) default to a generic relative
    risk/hazard-reduction heuristic (<5% / 5-20% / 20-40% / >40%), because GRADE
    does not define a universal numeric cutoff -- a real minimally-important-
    difference (MID) is outcome-specific and should be supplied by the review
    panel. Override the default bands with --mid-small/--mid-moderate/--mid-large
    if you have real MID-derived boundaries; there is no per-outcome override in
    this CLI (all outcomes in one run share the same bands) -- a documented
    limitation, not an oversight.
"""
import argparse, json, os, re, glob, sys, uuid, datetime

# ---------------------------------------------------------------------------
# GRADE plain-language templates (Cochrane Handbook Table 15.6.b)
# ---------------------------------------------------------------------------
CERTAINTY_LABEL = {"high": "high", "moderate": "moderate", "low": "low", "very_low": "very low"}

TEMPLATES = {
    "high":     {"large": "{I} results in a large {dirnoun} in {O}",
                 "moderate": "{I} {dirverb} {O}",
                 "small": "{I} {dirverb} {O} slightly",
                 "trivial": "{I} results in little to no difference in {O}"},
    "moderate": {"large": "{I} probably results in a large {dirnoun} in {O}",
                 "moderate": "{I} probably {dirverb} {O}",
                 "small": "{I} probably {dirverb} {O} slightly",
                 "trivial": "{I} probably results in little to no difference in {O}"},
    "low":      {"large": "{I} may result in a large {dirnoun} in {O}",
                 "moderate": "{I} may {dirverb} {O}",
                 "small": "{I} may {dirverb} {O} slightly",
                 "trivial": "{I} may result in little to no difference in {O}"},
}

MEASURE_TYPE_MAP = {"HR": "HazardRatio", "RR": "RiskRatio", "OR": "OddsRatio"}

DIRECTION_LABEL = {
    "INT_BETTER": "Intervention favourable",
    "COMP_BETTER": "Comparator favourable",
    "NO_DIFF": "No important difference",
    "UNCERTAIN": "High uncertainty",
}


def numify(s):
    """'420' -> 420 (int); '0.72' -> 0.72 (float)."""
    v = float(s)
    return int(v) if v == int(v) else v


def is_number(s):
    try:
        float(s); return True
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------
def build_colmap(ws, path):
    """Locate columns by header ROLE (row1 keyword match), not hardcoded letter --
    survives column reordering between exports."""
    row1 = {c: str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)}

    def find(keyword):
        for c, v in row1.items():
            if keyword.lower() in v.lower():
                return c
        return None

    cols = {"outcomes": find("outcome"), "relative": find("relative"),
            "absolute": find("absolute"), "certainty": find("certainty"),
            "citations": find("citation")}
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        raise SystemExit(f"{path}: could not find header column(s) for {missing} in row 1 "
                          f"(found: {row1}) -- unexpected Summary-of-Findings layout.")

    a = cols["absolute"]
    intervention_col, control_col, riskdiff_col = a, a + 1, a + 2
    row2 = {c: str(ws.cell(row=2, column=c).value or "") for c in range(a, min(a + 4, ws.max_column + 1))}
    for c, v in row2.items():
        lv = v.lower()
        if "intervention" in lv: intervention_col = c
        elif "control" in lv: control_col = c
        elif "difference" in lv: riskdiff_col = c

    return {"outcomes": cols["outcomes"], "relative": cols["relative"],
            "intervention": intervention_col, "control": control_col, "riskdiff": riskdiff_col,
            "certainty": cols["certainty"], "citations": cols["citations"]}


def find_data_start_row(ws):
    """Merged-header detection: header block is however many rows the row-1 merges span."""
    hr = 1
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)
    return hr + 1


def parse_excel_outcome(path):
    """Parse one outcome .xlsx -> dict of raw + derived fields. Raises SystemExit with
    an actionable message on any layout it doesn't recognize."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary of Findings"] if "Summary of Findings" in wb.sheetnames else wb.active

    colmap = build_colmap(ws, path)
    r0 = find_data_start_row(ws)
    if ws.cell(row=r0, column=colmap["outcomes"]).value is None:
        raise SystemExit(f"{path}: no data row found at row {r0} (below the detected header block).")

    def cell(col_key, row=r0):
        v = ws.cell(row=row, column=colmap[col_key]).value
        return "" if v is None else str(v).strip()

    # --- A: stratum label + stated citation count ---
    a_text = cell("outcomes")
    lines = [l.strip() for l in a_text.split("\n") if l.strip()]
    stated_n_citations = None
    if lines and re.match(r"^\d+\s+citations?$", lines[-1], re.I):
        stated_n_citations = int(re.match(r"^(\d+)", lines[-1]).group(1))
        lines = lines[:-1]
    stratum = " ".join(lines)

    # --- B: relative effect "PREFIX value\n(lo to hi)" ---
    b_text = cell("relative")
    m_point = re.search(r"([A-Za-z]{2,4})\s*([\d.]+)", b_text)
    if not m_point:
        raise SystemExit(f"{path}: could not parse relative-effect cell {b_text!r} "
                          f"(expected e.g. 'HR 0.72').")
    prefix, point = m_point.group(1).upper(), numify(m_point.group(2))
    m_ci = re.search(r"([\d.]+)\s*to\s*([\d.]+)", b_text, re.I) or \
           re.search(r"\(\s*([\d.]+)\s*[-–—]\s*([\d.]+)\s*\)", b_text)
    if not m_ci:
        raise SystemExit(f"{path}: could not parse the CI in relative-effect cell {b_text!r} "
                          f"(expected e.g. '(0.55 to 0.95)').")
    ci_lo, ci_hi = numify(m_ci.group(1)), numify(m_ci.group(2))
    if prefix not in MEASURE_TYPE_MAP:
        print(f"WARNING: {path}: unrecognized relative-effect prefix {prefix!r} "
              f"(expected HR/RR/OR) -- using it literally as the relativeEffect @type, "
              f"and skipping the absolute-effect CI transform (only implemented for HR/RR/OR).",
              file=sys.stderr)

    # --- C/D: arm absolute rates "N per 1000" ---
    def per1000(col_key):
        t = cell(col_key)
        m = re.search(r"([\d.]+)\s*per\s*1000", t, re.I)
        if not m:
            raise SystemExit(f"{path}: could not parse absolute-rate cell {t!r} (expected 'N per 1000').")
        return numify(m.group(1))

    intervention_rate = per1000("intervention")
    control_rate = per1000("control")

    # --- E: signed risk difference "N fewer/more per 1000" ---
    e_text = cell("riskdiff")
    m_rd = re.search(r"([\d.]+)\s*(fewer|less|more)\s*per\s*1000", e_text, re.I)
    if not m_rd:
        raise SystemExit(f"{path}: could not parse risk-difference cell {e_text!r} "
                          f"(expected e.g. '80 fewer per 1000').")
    rd_value = numify(m_rd.group(1))
    rd_signed = -rd_value if m_rd.group(2).lower() in ("fewer", "less") else rd_value

    # --- F: "Overall: <level>" certainty ---
    f_text = cell("certainty")
    m_cert = re.search(r"overall:?\s*(.+)", f_text, re.I)
    if not m_cert:
        raise SystemExit(f"{path}: could not parse certainty cell {f_text!r} (expected 'Overall: Moderate').")
    certainty_display = m_cert.group(1).strip().upper()          # e.g. "MODERATE", "VERY LOW"
    certainty_key = m_cert.group(1).strip().lower().replace(" ", "_")  # e.g. "moderate", "very_low"
    if certainty_key not in ("high", "moderate", "low", "very_low"):
        raise SystemExit(f"{path}: unrecognized certainty level {m_cert.group(1)!r} "
                          f"(expected High/Moderate/Low/Very low).")

    # --- G: citations, one per row from r0 downward ---
    citations = []
    for r in range(r0, ws.max_row + 1):
        v = ws.cell(row=r, column=colmap["citations"]).value
        if v is not None and str(v).strip():
            citations.append(str(v).strip())
    if stated_n_citations is not None and len(citations) != stated_n_citations:
        print(f"WARNING: {path}: outcome cell states {stated_n_citations} citations "
              f"but {len(citations)} citation row(s) were found -- using the found count.",
              file=sys.stderr)
    n_studies = len(citations) or stated_n_citations

    return {
        "path": path, "stratum": stratum, "prefix": prefix, "point": point,
        "ci_lo": ci_lo, "ci_hi": ci_hi, "intervention_rate": intervention_rate,
        "control_rate": control_rate, "rd_signed": rd_signed, "rd_text": e_text,
        "certainty_display": certainty_display, "certainty_key": certainty_key,
        "citations": citations, "n_studies": n_studies,
    }


def derive_outcome_name(path):
    """Filename -> display name: strip trailing _YYYYMMDD_HHMM, underscores -> spaces.
    'OS_(DrugX_vs_Placebo)_Overall_20260101_0900.xlsx' ->
    'OS (DrugX vs Placebo) Overall'."""
    stem = os.path.splitext(os.path.basename(path))[0]
    stem = re.sub(r"_\d{8}_\d{4}$", "", stem)
    return stem.replace("_", " ").strip()


def short_outcome_label(display_name):
    """For plain-language sentences: the leading metric token before any '(' --
    e.g. 'OS (DrugX vs Placebo) Overall' -> 'OS'. The source data never gives a
    spelled-out outcome name (see module docstring), so sentences necessarily use
    whatever abbreviation is in the filename."""
    return display_name.split("(")[0].strip() or display_name


# ---------------------------------------------------------------------------
# Relative-effect -> absolute-risk transform (for the absolute-effect CI)
# ---------------------------------------------------------------------------
def relative_to_risk(prefix, rel_value, control_risk_frac):
    """control_risk_frac and the return value are both fractions (0-1), not per-1000.
    Reverse-engineered from a live MAGICapp "Calculate estimates" run (see module
    docstring) -- confirmed to reproduce its output exactly for the HR case; RR/OR
    use the standard, textbook risk-ratio/odds-ratio transforms (not independently
    live-verified, but these two are the uncontroversial, universally-used forms)."""
    if prefix == "HR":
        return 1 - (1 - control_risk_frac) ** rel_value
    if prefix == "RR":
        return control_risk_frac * rel_value
    if prefix == "OR":
        control_odds = control_risk_frac / (1 - control_risk_frac)
        intervention_odds = control_odds * rel_value
        return intervention_odds / (1 + intervention_odds)
    return None  # unrecognized prefix -- caller must handle


def compute_absolute_effect_ci(prefix, ci_lo, ci_hi, control_rate_per1000):
    """Returns (ci_from_per1000_diff, ci_to_per1000_diff) -- the absolute-effect CI
    bounds AS A SIGNED DIFFERENCE FROM THE CONTROL RATE, per 1000, in the same
    source-preserved order as ci_lo/ci_hi (not re-sorted; mirrors how the relative
    effect's own CI bounds are kept in source order elsewhere in this script). Returns
    (None, None) for an unrecognized prefix -- nothing honest to compute."""
    control_frac = control_rate_per1000 / 1000.0
    risk_lo = relative_to_risk(prefix, ci_lo, control_frac)
    risk_hi = relative_to_risk(prefix, ci_hi, control_frac)
    if risk_lo is None or risk_hi is None:
        return None, None
    diff_lo = round((risk_lo - control_frac) * 1000, 2)
    diff_hi = round((risk_hi - control_frac) * 1000, 2)
    return diff_lo, diff_hi


# ---------------------------------------------------------------------------
# GRADE plain-language summary + direction-of-benefit classification
# (deterministic, no LLM -- one shared decision tree, see module docstring)
# ---------------------------------------------------------------------------
def classify_effect_magnitude(value, null, thresholds):
    """value = a point on the ratio scale (point estimate or a CI bound); null = 1.0
    for HR/RR/OR. Returns (magnitude, direction) where direction is 'reduction' or
    'increase' and magnitude in trivial/small/moderate/large, classified against the
    RELATIVE % change from null -- see module docstring re: MID vs. this default."""
    pct = (null - value) / null * 100.0   # positive = reduction, negative = increase
    direction = "reduction" if pct >= 0 else "increase"
    ap = abs(pct)
    t_small, t_mod, t_large = thresholds
    if ap < t_small: magnitude = "trivial"
    elif ap < t_mod: magnitude = "small"
    elif ap < t_large: magnitude = "moderate"
    else: magnitude = "large"
    return magnitude, direction


def crosses_null(ci_lo, ci_hi, null=1.0):
    return (ci_lo - null) * (ci_hi - null) < 0


def render_cell(certainty_key, magnitude, direction, intervention, outcome):
    dirverb = "reduces" if direction == "reduction" else "increases"
    dirnoun = "reduction" if direction == "reduction" else "increase"
    sentence = TEMPLATES[certainty_key][magnitude].format(I=intervention, O=outcome,
                                                            dirverb=dirverb, dirnoun=dirnoun)
    return sentence[0].upper() + sentence[1:]


def _direction_of_benefit(direction, favorable_direction):
    """direction = 'reduction'/'increase' (of the EVENT rate); favorable_direction =
    which of those is desirable (default 'reduction' -- see --favorable-direction)."""
    return "INT_BETTER" if direction == favorable_direction else "COMP_BETTER"


def generate_plain_language_summary(intervention, outcome_display, parsed, thresholds=(5, 20, 40),
                                     favorable_direction="reduction"):
    """Returns a dict: {"sentence": str, "flagged": bool, "direction_of_benefit": one of
    INT_BETTER/COMP_BETTER/NO_DIFF/UNCERTAIN}, per GRADE Table 15.6.b / Cochrane Handbook
    Ch.15 conventions, including the CI-crosses-null special cases (Handbook Sec 15.6.4).
    Both the sentence and the direction classification come from the same branching so
    they can never disagree with each other."""
    out = short_outcome_label(outcome_display)
    I, cert = intervention, parsed["certainty_key"]
    label = CERTAINTY_LABEL[cert]
    prefix, point, lo, hi = parsed["prefix"], parsed["point"], parsed["ci_lo"], parsed["ci_hi"]
    rd_text = parsed["rd_text"]
    stats_ci = f"{prefix} {point}, 95% CI {lo} to {hi}"

    if cert == "very_low":
        sentence = f"The evidence is very uncertain about the effect of {I} on {out}"
        return {"sentence": f"{sentence} ({stats_ci}; {label}-certainty evidence).",
                "flagged": False, "direction_of_benefit": "UNCERTAIN"}

    if not crosses_null(lo, hi):
        magnitude, direction = classify_effect_magnitude(point, 1.0, thresholds)
        sentence = render_cell(cert, magnitude, direction, I, out)
        dob = "NO_DIFF" if magnitude == "trivial" else _direction_of_benefit(direction, favorable_direction)
        return {"sentence": f"{sentence} ({stats_ci}; {rd_text}; {label}-certainty evidence).",
                "flagged": False, "direction_of_benefit": dob}

    lo_mag, lo_dir = classify_effect_magnitude(lo, 1.0, thresholds)
    hi_mag, hi_dir = classify_effect_magnitude(hi, 1.0, thresholds)

    if lo_mag == "trivial" and hi_mag == "trivial":
        sentence = render_cell(cert, "trivial", "reduction", I, out)
        return {"sentence": f"{sentence} ({stats_ci}; {rd_text}; {label}-certainty evidence).",
                "flagged": False, "direction_of_benefit": "NO_DIFF"}

    if lo_mag == "trivial" or hi_mag == "trivial":
        important_dir = hi_dir if lo_mag == "trivial" else lo_dir
        important_verb = "reduce" if important_dir == "reduction" else "increase"
        base = render_cell(cert, "trivial", "reduction", I, out)
        sentence = f"{base}, but may {important_verb} {out}"
        return {"sentence": f"{sentence} ({stats_ci}; {rd_text}; {label}-certainty evidence).",
                "flagged": False, "direction_of_benefit": "UNCERTAIN"}

    # both bounds important, opposite directions -> direction genuinely undetermined
    point_dir = "reduction" if point <= 1 else "increase"
    point_verb = "reduce" if point_dir == "reduction" else "increase"
    opp_verb = "increase" if point_dir == "reduction" else "reduce"
    sentence = (f"{I} may {point_verb} {out}, but it may also {opp_verb} it — the evidence "
                f"does not allow a clear conclusion about the direction of the effect")
    stats = (f"{stats_ci}; approximately {rd_text} at the point estimate, but the interval is "
             f"compatible with either a large reduction or an increase in events; "
             f"{label}-certainty evidence")
    return {"sentence": f"{sentence} ({stats}).", "flagged": True, "direction_of_benefit": "UNCERTAIN"}


# ---------------------------------------------------------------------------
# JSON-LD assembly
# ---------------------------------------------------------------------------
GDT_CONTEXT = [
    {"@base": "http://dbep.gradepro.org/", "@vocab": "http://dbep.gradepro.org/schema/",
     "@language": "en", "schema": "http://schema.org/",
     "cochrane": "http://linkeddata.cochrane.org/ontologies/pico/",
     "xsd": "http://www.w3.org/2001/XMLSchema#", "version": "schema:version", "name": "schema:name",
     "sameAs": "schema:sameAs",
     "publicationTime": {"@id": "schema:datePublished", "@type": "xsd:dateTime"},
     "modificationTime": {"@id": "schema:dateModified", "@type": "xsd:dateTime"},
     "author": {"@id": "schema:author", "@container": "@list"},
     "footnote": {"@container": "@list"}, "code": "schema:code", "codeValue": "schema:codeValue",
     "codingSystem": "schema:codingSystem", "unitCode": "schema:unitCode", "unitText": "schema:unitText",
     "value": "schema:value", "text": "schema:value", "lowerBound": "schema:minValue",
     "upperBound": "schema:maxValue", "Person": "schema:Person", "MedicalCode": "schema:MedicalCode",
     "MedicalSymptom": "schema:MedicalSymptom", "MedicalCondition": "schema:MedicalCondition",
     "QuantitativeValue": "schema:QuantitativeValue"},
    {"PICO": "cochrane:PICO", "healthProblemOrPopulation": "cochrane:population",
     "condition": "cochrane:condition", "intervention": "cochrane:intervention",
     "comparison": "cochrane:comparison", "outcome": {"@id": "cochrane:outcome", "@container": "@list"}},
    {"title": "name", "sofTitle": "alternateName",
     "evidenceSummary": {"@container": "@list"}, "patientGroup": {"@container": "@list"}},
]


def build_outcome_entry(outcome_id, display_name):
    return {"@id": outcome_id, "@type": ["Outcome", "TimeToEvent"],
            "name": {"value": display_name}, "event": {"@type": "Event"}}


def build_evidence_summary_entry(outcome_id, parsed):
    measure_type = MEASURE_TYPE_MAP.get(parsed["prefix"], parsed["prefix"])
    ci_from, ci_to = compute_absolute_effect_ci(parsed["prefix"], parsed["ci_lo"], parsed["ci_hi"],
                                                 parsed["control_rate"])
    return {
        "@type": "DichotomousData",
        "forOutcome": {"@id": outcome_id},
        "studyDesign": {"@type": "RandomisedTrials", "name": "randomised trials"},
        "numberOfStudies": {"value": parsed["n_studies"]},
        "patientGroup": [
            {"@type": "InterventionGroup", "totalCount": {"value": None}},
            {"@type": "ControlGroup", "totalCount": {"value": None}},
        ],
        "measuredWith": {"@type": "OutcomeMeasure", "name": ""},
        "effectSummary": {
            "@type": "Pooled",
            "relativeEffect": {"@type": measure_type, "value": {"value": parsed["point"]},
                                "confidenceLevel": {"value": 0.95},
                                "confidenceIntervalFrom": parsed["ci_lo"],
                                "confidenceIntervalTo": parsed["ci_hi"]},
            "absoluteEffect": [{
                "@type": "AutoCalculatedAbsoluteEffect", "forControlRisk": {"@id": "_:cr1"},
                "value": {"value": parsed["rd_signed"]}, "confidenceLevel": {"value": 0.95},
                "confidenceIntervalFrom": ci_from, "confidenceIntervalTo": ci_to, "denominator": 1000,
            }],
        },
        "quality": {"@type": "GradeQuality", "value": None, "name": parsed["certainty_display"],
                     "riskOfBias": None, "inconsistency": None, "indirectness": None, "imprecision": None,
                     "otherConsiderations": {"name": "", "publicationBias": None,
                        "doseResponseGradient": {"@type": "NoChange", "name": "no"},
                        "plausibleConfounding": {"@type": "NoChange", "name": "no"},
                        "largeEffect": {"@type": "NoChange", "name": "no"}}},
        "controlRisk": [{"@id": "_:cr1", "@type": "LowControlRisk",
                          "value": round(parsed["control_rate"] / 10.0, 4)}],
    }


def assemble_jsonld(outcomes, population, intervention, comparator, title, sof_title):
    pico_id = str(uuid.uuid4())
    question_outcomes, evidence_summaries, explanations = [], [], [
        {"@type": "Explanation", "@id": "_:e0", "text": "no_explanation_provided"},
    ]
    for i, (display_name, parsed, pls) in enumerate(outcomes):
        outcome_id = f"outcomes/{uuid.uuid4()}"
        question_outcomes.append(build_outcome_entry(outcome_id, display_name))
        evidence_summaries.append(build_evidence_summary_entry(outcome_id, parsed))
        if pls is not None:
            explanations.append({"@type": "Explanation", "@id": f"_:pls{i}", "text": pls["sentence"],
                                  "forOutcome": {"@id": outcome_id}})

    default_title = f"Should {intervention} vs. {comparator} be used for {population}?"
    default_sof_title = f"{intervention} vs. {comparator} for {population}"

    return {
        "@id": pico_id,
        "@context": GDT_CONTEXT,
        "@type": "ManagementEvidenceProfile",
        "version": "1",
        "modificationTime": datetime.datetime.now().astimezone().isoformat(timespec="seconds"),
        "bibliography": {"value": ""},
        "title": title or default_title,
        "sofTitle": sof_title or default_sof_title,
        "question": {
            "@id": f"questions/{pico_id}", "@type": "PICO",
            "healthProblemOrPopulation": {"@type": "HealthProblemOrPopulation",
                                           "name": {"value": population}, "setting": {"value": ""}},
            "intervention": {"name": intervention, "@type": "Intervention"},
            "comparison": {"name": comparator, "@type": "Comparison"},
            "outcome": question_outcomes,
        },
        "explanation": explanations,
        "evidenceSummary": evidence_summaries,
        "reference": [],
    }


def write_post_import_checklist_md(path, intervention, comparator, population, rows, flagged_count):
    with open(path, "w") as f:
        f.write(f"# Post-import checklist: {intervention} vs. {comparator}\n\n")
        f.write(f"Population: {population}\n\n")
        f.write(
            "MAGICapp's beta GDT importer correctly imports the relative effect and control-arm risk "
            "for every outcome below, but confirmed live testing found it leaves three things blank/unset "
            "that it cannot currently be made to fill in from the JSON-LD file alone. For **each** outcome, "
            "after import:\n\n"
            "1. Open the outcome's **Absolute effect estimates** cell (pencil icon) and click "
            "**\"Calculate estimates\"** (or check \"Auto-calculated\") -- this computes the intervention-arm "
            "rate from the control risk + relative effect that *did* import correctly. The expected value is "
            "given below so you can confirm it, not guess at it.\n"
            "2. Set **\"Direction of benefit\"** (in the same edit panel, under Certainty of the evidence) to "
            "the recommended value below.\n"
            "3. Paste the **plain-language summary** sentence below into that outcome's \"Plain language "
            "summary\" field.\n\n"
        )
        if flagged_count:
            f.write(f"**{flagged_count} outcome(s) below are flagged `[NEEDS HUMAN REVIEW]`** -- "
                    "their confidence interval crosses the null with an important effect plausible "
                    "in BOTH directions, so GRADE's own convention is to state direction is "
                    "undetermined rather than default to the point estimate. Review before using.\n\n")
        f.write("---\n\n")
        for display_name, parsed, pls in rows:
            tag = " `[NEEDS HUMAN REVIEW]`" if pls["flagged"] else ""
            f.write(f"## {display_name}{tag}\n\n")
            f.write(f"- **Expected intervention-arm rate after \"Calculate estimates\":** "
                    f"~{parsed['intervention_rate']} per 1000 (this script's parse of your source Excel's "
                    f"own stated value -- MAGICapp's own calculation may land within a point or two of this "
                    f"due to its transform formula's rounding; if it's off by more than that, something's "
                    f"wrong, don't just accept it).\n")
            f.write(f"- **Direction of benefit:** {DIRECTION_LABEL[pls['direction_of_benefit']]}\n")
            f.write(f"- **Plain language summary:**\n\n  {pls['sentence']}\n\n")


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="one outcome's .xlsx, or a folder of many (one outcome each)")
    ap.add_argument("--population", required=True)
    ap.add_argument("--intervention", required=True)
    ap.add_argument("--comparator", required=True)
    ap.add_argument("--title", default=None)
    ap.add_argument("--sof-title", default=None)
    ap.add_argument("--outcome-name", default=None,
                     help="override the filename-derived outcome name (single-file input only)")
    ap.add_argument("--mid-small", type=float, default=5.0,
                     help="trivial/small relative-%% boundary (default 5, non-authoritative default -- see docstring)")
    ap.add_argument("--mid-moderate", type=float, default=20.0, help="small/moderate boundary (default 20)")
    ap.add_argument("--mid-large", type=float, default=40.0, help="moderate/large boundary (default 40)")
    ap.add_argument("--favorable-direction", choices=["reduction", "increase"], default="reduction",
                     help="which direction of the EVENT rate is desirable, for the Direction-of-benefit "
                          "classification (default 'reduction' -- correct for OS/DFS/PFS/RFS-style outcomes "
                          "where the event is death/progression/recurrence; see docstring)")
    ap.add_argument("--no-pls", action="store_true",
                     help="skip plain-language summary + direction-of-benefit generation")
    ap.add_argument("-o", "--output", required=True, help="output .json path")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        ap.error(f"input path not found: {args.input}")

    if os.path.isdir(args.input):
        if args.outcome_name:
            ap.error("--outcome-name only applies to single-file input, not a folder")
        paths = sorted(p for p in glob.glob(os.path.join(args.input, "*.xlsx"))
                        if not os.path.basename(p).startswith(("~$", ".")))
        if not paths:
            ap.error(f"no .xlsx files found in folder: {args.input}")
    else:
        paths = [args.input]

    thresholds = (args.mid_small, args.mid_moderate, args.mid_large)
    outcomes, checklist_rows, flagged_count = [], [], 0
    for path in paths:
        parsed = parse_excel_outcome(path)
        display_name = args.outcome_name if (args.outcome_name and len(paths) == 1) else derive_outcome_name(path)
        pls = None
        if not args.no_pls:
            pls = generate_plain_language_summary(args.intervention, display_name, parsed, thresholds,
                                                    args.favorable_direction)
            if pls["flagged"]:
                flagged_count += 1
            checklist_rows.append((display_name, parsed, pls))
        outcomes.append((display_name, parsed, pls))

    doc = assemble_jsonld(outcomes, args.population, args.intervention, args.comparator,
                           args.title, args.sof_title)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)) or ".", exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(doc, f, indent=1)

    print(f"Wrote {args.output} ({len(outcomes)} outcome(s), {len(paths)} source file(s)).")

    if not args.no_pls:
        checklist_path = os.path.splitext(args.output)[0] + "_post_import_checklist.md"
        write_post_import_checklist_md(checklist_path, args.intervention, args.comparator, args.population,
                                        checklist_rows, flagged_count)
        print(f"Wrote {checklist_path}.")
        if flagged_count:
            print(f"  {flagged_count} outcome(s) flagged NEEDS HUMAN REVIEW (CI crosses null with "
                  f"important effect plausible either direction) -- see the checklist file.")

    print("Reminder: this is a beta MAGICapp import path. After importing, work through the post-import "
          "checklist for each outcome (Calculate estimates, Direction of benefit, plain-language summary) "
          "-- confirmed live that MAGICapp's GDT importer does not set any of these three on its own, even "
          "though it does correctly import the relative effect and control-arm risk needed to compute the "
          "first one. patientGroup N per arm, study design, and GRADE domain sub-ratings are also left "
          "null/default here -- see SKILL.md for the full list.")


if __name__ == "__main__":
    main()
