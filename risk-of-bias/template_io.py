"""
Template I/O for the Living Perioperative RCC risk-of-bias workbooks.

This module renders a RoB 2 assessment (an ``assess_rob2()`` result dict from
``kernel.py``) into the exact layout, vocabulary and cells of the project's
reviewer templates (``rob_reviewer_1.xlsx`` / ``rob_reviewer_2.xlsx``) and writes
it *into* the real workbook, preserving every other sheet, the seeded study rows,
and the data-validation dropdowns.

Template layout (one row per assessed result; one sheet per outcome)
--------------------------------------------------------------------
  A paper_id | B pubmed_id | C nct_number | D outcome_id | E outcome_name
  form_data:      F comparator | G experimental | H weight_analysis |
                  I numerical_result | J aim | K sources | L adhering-preamble
  domain_records: d1q1..d1q3, d2q1..d2q7, d3q1..d3q4, d4q1..d4q5, d5q1..d5q3
                  (each answer column is followed by a 'description' column)
  domain_scores:  per domain D1..D5 -> (direction_of_bias, assessor_judgement)
  overall_score:  assessor_judgement (BO) | overall_description (BP)

The signalling-question id ``d{n}q{m}`` maps positionally to RoB 2 question
``{n}.{m}`` (d1q1 = 1.1, ... d5q3 = 5.3).

Project conventions (locked with the user)
------------------------------------------
* Answer vocabulary (dropdown): Yes / Probably Yes / Probably No / No /
  No Information, plus NA for questions the flowchart does not reach.
* Judgement vocabulary (dropdown): Low risk / Some concerns / High risk.
* Direction of bias (dropdown): Favours experimental / Favours comparator /
  No direction.
* By default we MATCH the completed examples: write the signalling answers, the
  five per-domain ``assessor_judgement`` cells and the overall ``assessor_judgement``
  cell; leave every ``description`` cell and ``direction_of_bias`` blank.

The column map is rebuilt from the header rows (1-3) at run time, so it stays
correct even if a sheet's columns shift.
"""
import openpyxl
from openpyxl.utils import get_column_letter as _CL

# --------------------------------------------------------------------------- #
# Vocabularies (canonical kernel value -> template cell text)
# --------------------------------------------------------------------------- #
TEMPLATE_ANSWER = {
    "Y": "Yes", "PY": "Probably Yes", "PN": "Probably No",
    "N": "No", "NI": "No Information", "NA": "NA",
}
TEMPLATE_JUDGEMENT = {
    "low": "Low risk", "some concerns": "Some concerns", "high": "High risk",
}
TEMPLATE_AIM = {
    "assignment": "Effect of assignment to intervention",
    "adhering": "Effect of adhering to intervention",
}
TEMPLATE_DIRECTION = {
    "experimental": "Favours experimental",
    "comparator": "Favours comparator",
    "none": "No direction", "no direction": "No direction", None: None,
}

# RoB 2 signalling questions per domain, in the template's positional order.
DOMAIN_QIDS = {
    "D1": ["1.1", "1.2", "1.3"],
    "D2": ["2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7"],
    "D3": ["3.1", "3.2", "3.3", "3.4"],
    "D4": ["4.1", "4.2", "4.3", "4.4", "4.5"],
    "D5": ["5.1", "5.2", "5.3"],
}
DOMAIN_ORDER = ["D1", "D2", "D3", "D4", "D5"]
HEADER_ROWS = 3          # rows 1-3 are the header block
FIRST_DATA_ROW = 4       # study rows begin at row 4


def _norm(s):
    return "" if s is None else str(s).strip()


# --------------------------------------------------------------------------- #
# Column map (rebuilt from the header block, robust to column shifts)
# --------------------------------------------------------------------------- #
def build_column_map(ws):
    """Scan header rows 1-3 of a worksheet and return a column map:

      {
        'meta':   {'paper_id':1,'pubmed_id':2,'nct_number':3,'outcome_id':4,'outcome_name':5},
        'form':   {'comparator':c,'experimental':c,'weight_analysis':c,
                   'numerical_result':c,'aim':c,'sources':c,'adhering_preamble':c},
        'q':      {'1.1':c, ...},          # answer columns
        'qdesc':  {'1.1':c, ...},          # the 'description' column after each answer
        'dir':    {'D1':c, ...},           # per-domain direction_of_bias
        'judg':   {'D1':c, ...},           # per-domain assessor_judgement
        'overall_judg': c, 'overall_desc': c,
      }
    """
    maxc = ws.max_column
    r1 = {c: _norm(ws.cell(1, c).value) for c in range(1, maxc + 1)}
    r2 = {c: _norm(ws.cell(2, c).value) for c in range(1, maxc + 1)}
    r3 = {c: _norm(ws.cell(3, c).value) for c in range(1, maxc + 1)}
    # forward-fill the row-1 group label across its columns, and the row-2 sub-label
    # WITHIN each row-1 group (the domain label sits only on the merged direction
    # column, so BF/BH/... inherit their 'domain_N' from BE/BG/...).
    grp, g = {}, ""
    mid_fill_map, mid_fill = {}, ""
    for c in range(1, maxc + 1):
        if r1[c]:
            if r1[c] != g:
                mid_fill = ""            # reset sub-label at each new group
            g = r1[c]
        grp[c] = g
        if r2[c]:
            mid_fill = r2[c]
        mid_fill_map[c] = mid_fill

    m = {"meta": {}, "form": {}, "q": {}, "qdesc": {},
         "dir": {}, "judg": {}, "overall_judg": None, "overall_desc": None}

    for c in range(1, maxc + 1):
        top, mid, lab = r1[c], mid_fill_map[c], r3[c]
        # top-level identity columns
        if top in ("paper_id", "pubmed_id", "nct_number", "outcome_id", "outcome_name"):
            m["meta"][top] = c
            continue
        group = grp[c]
        if group == "form_data":
            if lab == "comparator":
                m["form"]["comparator"] = c
            elif lab == "experimental":
                m["form"]["experimental"] = c
            elif lab == "weight_analysis":
                m["form"]["weight_analysis"] = c
            elif lab == "numerical_result":
                m["form"]["numerical_result"] = c
            elif lab.lower().startswith("is the review team"):
                m["form"]["aim"] = c
            elif lab.lower().startswith("which of the following sources"):
                m["form"]["sources"] = c
            elif lab.lower().startswith("if the aim is to assess"):
                m["form"]["adhering_preamble"] = c
        elif group == "domain_records":
            if lab.startswith("d") and "q" in lab:
                # dNqM -> N.M
                dnum = lab[1:lab.index("q")]
                qnum = lab[lab.index("q") + 1:]
                qid = "%s.%s" % (dnum, qnum)
                m["q"][qid] = c
                # the immediately following column is this question's description
                if c + 1 <= maxc and r3.get(c + 1, "").lower() == "description":
                    m["qdesc"][qid] = c + 1
        elif group == "domain_scores":
            dom = "D" + mid[-1] if mid.startswith("domain_") else None
            if dom:
                if lab == "direction_of_bias":
                    m["dir"][dom] = c
                elif lab == "assessor_judgement":
                    m["judg"][dom] = c
        elif group == "overall_score":
            if lab == "assessor_judgement":
                m["overall_judg"] = c
            elif lab == "overall_description":
                m["overall_desc"] = c
    return m


# --------------------------------------------------------------------------- #
# Render an assess_rob2() result into template values
# --------------------------------------------------------------------------- #
def assessment_to_template_values(result, comparator=None, experimental=None,
                                  effect=None, na_for_unreached=True):
    """Convert an ``assess_rob2()`` result dict into template-vocabulary values.

    Returns a dict::

        {'answers': {'1.1':'Yes', ...},           # every template question, NA-filled
         'domain_judgements': {'D1':'Low risk', ...},
         'overall': 'Low risk',
         'aim': 'Effect of assignment to intervention',
         'comparator': ..., 'experimental': ...}

    Unreached questions become 'NA' (matching the completed examples) unless
    ``na_for_unreached=False``, in which case they are left as '' (blank).
    """
    eff = effect or result.get("effect") or "assignment"
    domains = result.get("domains", {})
    answers = {}
    for dom in DOMAIN_ORDER:
        given = (domains.get(dom, {}) or {}).get("answers", {}) or {}
        for qid in DOMAIN_QIDS[dom]:
            code = given.get(qid)
            if code in (None, ""):
                answers[qid] = "NA" if na_for_unreached else ""
            else:
                answers[qid] = TEMPLATE_ANSWER.get(str(code).upper(), str(code))
    dj = {}
    for dom in DOMAIN_ORDER:
        j = (domains.get(dom, {}) or {}).get("judgement") or result.get("domain_judgements", {}).get(dom)
        dj[dom] = TEMPLATE_JUDGEMENT.get(_norm(j).lower(), _norm(j))
    overall = TEMPLATE_JUDGEMENT.get(_norm(result.get("overall")).lower(), _norm(result.get("overall")))
    return {
        "answers": answers,
        "domain_judgements": dj,
        "overall": overall,
        "aim": TEMPLATE_AIM.get(eff, TEMPLATE_AIM["assignment"]),
        "comparator": comparator if comparator is not None else result.get("comparator"),
        "experimental": experimental if experimental is not None else result.get("experimental"),
    }


# --------------------------------------------------------------------------- #
# Locate the row for a study within a sheet
# --------------------------------------------------------------------------- #
def find_row(ws, colmap, paper_id, outcome_id=None):
    """Return the 1-based row whose paper_id (and optional outcome_id) matches, or None.

    The reviewer templates pre-seed one row per study on every outcome sheet, so a
    matching row already exists; we fill it rather than appending.
    """
    pcol = colmap["meta"]["paper_id"]
    ocol = colmap["meta"].get("outcome_id")
    pid = _norm(paper_id)
    oid = _norm(outcome_id) if outcome_id is not None else None
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if _norm(ws.cell(r, pcol).value) == pid:
            if oid is None or ocol is None or _norm(ws.cell(r, ocol).value) == oid:
                return r
    return None


# --------------------------------------------------------------------------- #
# Write one assessment into an already-open worksheet
# --------------------------------------------------------------------------- #
def write_assessment_ws(ws, paper_id, result, comparator=None, experimental=None,
                        effect=None, outcome_id=None, colmap=None,
                        write_descriptions=False, write_direction=False,
                        overall_description=None, na_for_unreached=True,
                        create_row_if_missing=False):
    """Write one RoB 2 result into worksheet ``ws`` for ``paper_id``.

    By default (matching the completed examples) writes signalling answers, the five
    per-domain judgements and the overall judgement; leaves description cells and
    direction_of_bias blank. Set ``write_descriptions``/``write_direction``/
    ``overall_description`` to populate those too.

    Returns the row written, or None if the study row was not found and
    ``create_row_if_missing`` is False.
    """
    colmap = colmap or build_column_map(ws)
    row = find_row(ws, colmap, paper_id, outcome_id)
    if row is None:
        if not create_row_if_missing:
            return None
        row = ws.max_row + 1
        ws.cell(row, colmap["meta"]["paper_id"]).value = paper_id
        if outcome_id is not None and colmap["meta"].get("outcome_id"):
            ws.cell(row, colmap["meta"]["outcome_id"]).value = outcome_id

    vals = assessment_to_template_values(result, comparator, experimental, effect,
                                         na_for_unreached=na_for_unreached)
    # form_data
    if colmap["form"].get("aim"):
        ws.cell(row, colmap["form"]["aim"]).value = vals["aim"]
    if vals["comparator"] is not None and colmap["form"].get("comparator"):
        ws.cell(row, colmap["form"]["comparator"]).value = vals["comparator"]
    if vals["experimental"] is not None and colmap["form"].get("experimental"):
        ws.cell(row, colmap["form"]["experimental"]).value = vals["experimental"]
    # signalling answers (+ optional descriptions)
    given_desc = {}
    for dom in DOMAIN_ORDER:
        d = (result.get("domains", {}).get(dom, {}) or {})
        for qid in DOMAIN_QIDS[dom]:
            c = colmap["q"].get(qid)
            if c:
                ws.cell(row, c).value = vals["answers"][qid]
        if write_descriptions:
            given_desc[dom] = d.get("evidence") or d.get("rationale")
    if write_descriptions:
        # put the domain-level rationale on that domain's first question description cell
        for dom in DOMAIN_ORDER:
            txt = given_desc.get(dom)
            if txt:
                q0 = DOMAIN_QIDS[dom][0]
                dc = colmap["qdesc"].get(q0)
                if dc:
                    ws.cell(row, dc).value = txt
    # per-domain judgements (+ optional direction)
    for dom in DOMAIN_ORDER:
        jc = colmap["judg"].get(dom)
        if jc:
            ws.cell(row, jc).value = vals["domain_judgements"][dom]
        if write_direction:
            dinfo = (result.get("domains", {}).get(dom, {}) or {}).get("direction")
            dc = colmap["dir"].get(dom)
            if dc and dinfo:
                ws.cell(row, dc).value = TEMPLATE_DIRECTION.get(str(dinfo).lower(), dinfo)
    # overall
    if colmap["overall_judg"]:
        ws.cell(row, colmap["overall_judg"]).value = vals["overall"]
    if overall_description and colmap["overall_desc"]:
        ws.cell(row, colmap["overall_desc"]).value = overall_description
    return row


# --------------------------------------------------------------------------- #
# Write a whole study (many outcome sheets) into one reviewer workbook
# --------------------------------------------------------------------------- #
def write_study(xlsx_path, paper_id, per_outcome, comparator=None, experimental=None,
                effect="assignment", out_path=None, **kw):
    """Fill one reviewer workbook for one study across its assessed outcome sheets.

    ``per_outcome`` maps a sheet name -> an ``assess_rob2()`` result (or a dict
    ``{'result':..., 'comparator':..., 'experimental':..., 'effect':...,
    'outcome_id':..., 'overall_description':...}`` to override per sheet).
    Only the sheets you pass are written; every other sheet/row is untouched, and the
    data-validation dropdowns are preserved. Saves to ``out_path`` (default: in place).

    Returns {sheet_name: row_written_or_None}.
    """
    wb = openpyxl.load_workbook(xlsx_path)          # keep formulas + validations
    written = {}
    for sheet, spec in per_outcome.items():
        if sheet not in wb.sheetnames:
            written[sheet] = None
            continue
        ws = wb[sheet]
        if isinstance(spec, dict) and "result" in spec:
            res = spec["result"]
            row = write_assessment_ws(
                ws, paper_id, res,
                comparator=spec.get("comparator", comparator),
                experimental=spec.get("experimental", experimental),
                effect=spec.get("effect", effect),
                outcome_id=spec.get("outcome_id"),
                overall_description=spec.get("overall_description"),
                **kw)
        else:
            row = write_assessment_ws(ws, paper_id, spec, comparator=comparator,
                                      experimental=experimental, effect=effect, **kw)
        written[sheet] = row
    wb.save(out_path or xlsx_path)
    return written


# --------------------------------------------------------------------------- #
# Read a template row back into canonical answers (for tests / adjudication)
# --------------------------------------------------------------------------- #
_ANSWER_BACK = {v.lower(): k for k, v in TEMPLATE_ANSWER.items()}
_JUDGE_BACK = {v.lower(): k for k, v in TEMPLATE_JUDGEMENT.items()}


def read_row(ws, paper_id, colmap=None, outcome_id=None):
    """Read a filled template row back into canonical kernel form:

        {'answers_by_domain': {'D1':{'1.1':'Y',...}, ...},
         'domain_judgements': {'D1':'low', ...},
         'overall': 'low', 'aim': '...', 'comparator':..., 'experimental':...}

    Cells that are blank/NA are omitted from the per-domain answer dicts. Useful to
    round-trip a completed example and re-run the kernel over it.
    """
    colmap = colmap or build_column_map(ws)
    row = find_row(ws, colmap, paper_id, outcome_id)
    if row is None:
        return None
    abd = {}
    for dom in DOMAIN_ORDER:
        d = {}
        for qid in DOMAIN_QIDS[dom]:
            c = colmap["q"].get(qid)
            v = _norm(ws.cell(row, c).value) if c else ""
            if v and v.upper() != "NA":
                d[qid] = _ANSWER_BACK.get(v.lower(), v)
        abd[dom] = d
    dj = {}
    for dom in DOMAIN_ORDER:
        c = colmap["judg"].get(dom)
        v = _norm(ws.cell(row, c).value) if c else ""
        dj[dom] = _JUDGE_BACK.get(v.lower(), v) if v else None
    ov = _norm(ws.cell(row, colmap["overall_judg"]).value) if colmap["overall_judg"] else ""
    return {
        "row": row,
        "answers_by_domain": abd,
        "domain_judgements": dj,
        "overall": _JUDGE_BACK.get(ov.lower(), ov) if ov else None,
        "aim": _norm(ws.cell(row, colmap["form"]["aim"]).value) if colmap["form"].get("aim") else None,
        "comparator": _norm(ws.cell(row, colmap["form"]["comparator"]).value) if colmap["form"].get("comparator") else None,
        "experimental": _norm(ws.cell(row, colmap["form"]["experimental"]).value) if colmap["form"].get("experimental") else None,
    }


def list_seeded_studies(xlsx_path, sheet="OS"):
    """Return [(row, paper_id, outcome_id, outcome_name)] seeded on a sheet (default OS)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]
    cm = build_column_map(ws)
    out = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        pid = ws.cell(r, cm["meta"]["paper_id"]).value
        if pid not in (None, ""):
            out.append((r, pid,
                        ws.cell(r, cm["meta"].get("outcome_id", 4)).value,
                        ws.cell(r, cm["meta"].get("outcome_name", 5)).value))
    return out
