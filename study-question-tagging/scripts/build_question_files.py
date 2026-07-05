#!/usr/bin/env python3
"""Build one Excel file per clinical question from an assignment matrix.

The deliverable for each question is the included-studies sheet filtered to just the
studies assigned to that question — same columns, same formatting. We get that by
copying the source workbook and deleting the rows that don't belong, which preserves
the header styling, column widths, and frozen panes exactly (rewriting from scratch
would lose them). A question with no assigned studies yields a header-only file, which
is the honest representation of "no included study answers this."

Usage:
  python build_question_files.py --source studies.xlsx --matrix matrix.json --out-dir DIR
                                 [--sheet NAME] [--id-col LETTER_OR_HEADER]

matrix.json (rich form, preferred):
  {
    "questions": [
      {"key": "CQ10", "slug": "premenopausal_ET_treatment",
       "title": "optional human title", "study_ids": ["12301","30503"]},
      ...
    ]
  }
matrix.json (flat form, also accepted):  {"CQ10": ["12301","30503"], "CQ11": [...]}

Filename per question = "<key>_<slug>.xlsx" (or "<key>.xlsx" if no slug), unless an
explicit "filename" is given on the question object.
"""
import argparse, json, os, re, shutil, sys


def col_to_idx(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_id_col(headers, override):
    if override:
        ov = override.strip()
        if re.fullmatch(r"[A-Za-z]{1,3}", ov) and ov.upper() == ov:
            return col_to_idx(ov)
        for i, h in enumerate(headers, 1):
            if h and str(h).strip().lower() == ov.lower():
                return i
        sys.exit(f"--id-col '{override}' not found in headers")
    for i, h in enumerate(headers, 1):
        if h and str(h).strip().lower() in ("id", "study id", "paper id", "record id"):
            return i
    return 1


def normalize_matrix(m):
    """Return a list of {key, slug, title, filename, study_ids}."""
    if isinstance(m, dict) and "questions" in m:
        qs = m["questions"]
    elif isinstance(m, dict):
        qs = [{"key": k, "study_ids": v} for k, v in m.items()]
    else:
        sys.exit("matrix.json must be an object")
    out = []
    for q in qs:
        key = str(q["key"]).strip()
        slug = q.get("slug")
        fn = q.get("filename")
        if not fn:
            safe_key = re.sub(r"[^A-Za-z0-9._-]+", "_", key)
            fn = f"{safe_key}_{slug}.xlsx" if slug else f"{safe_key}.xlsx"
        ids = [str(x).strip() for x in (q.get("study_ids") or [])]
        out.append({"key": key, "slug": slug, "title": q.get("title"),
                    "filename": fn, "study_ids": ids})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True)
    ap.add_argument("--matrix", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--sheet")
    ap.add_argument("--id-col")
    a = ap.parse_args()

    import openpyxl
    os.makedirs(a.out_dir, exist_ok=True)
    with open(a.matrix) as f:
        questions = normalize_matrix(json.load(f))

    wb0 = openpyxl.load_workbook(a.source, data_only=True)
    sheet_name = a.sheet or wb0.sheetnames[0]
    ws0 = wb0[sheet_name]
    headers = [ws0.cell(row=1, column=c).value for c in range(1, ws0.max_column + 1)]
    id_idx = resolve_id_col(headers, a.id_col)

    id_by_row = {}
    for r in range(2, ws0.max_row + 1):
        v = ws0.cell(row=r, column=id_idx).value
        if v is not None and str(v).strip():
            id_by_row[r] = str(v).strip()
    all_ids = set(id_by_row.values())

    summary = []
    for q in questions:
        keep = set(q["study_ids"])
        missing = keep - all_ids
        if missing:
            sys.exit(f"{q['key']}: IDs not in source sheet: {sorted(missing)}")

        out_path = os.path.join(a.out_dir, q["filename"])
        shutil.copyfile(a.source, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb[sheet_name]
        for r in sorted(id_by_row.keys(), reverse=True):
            if id_by_row[r] not in keep:
                ws.delete_rows(r, 1)
        wb.save(out_path)
        summary.append((q["key"], q["filename"], len(keep), sorted(keep)))

    print("Built question files in", a.out_dir)
    assigned = set()
    for key, fn, n, ids in summary:
        print(f"  {key:8} {n:3d} studies -> {fn}")
        print(f"           {ids}")
        assigned |= set(ids)
    unassigned = sorted(all_ids - assigned)
    if unassigned:
        print(f"\n[!] {len(unassigned)} source studies are in NO question file "
              f"(confirm this is intended): {unassigned}")


if __name__ == "__main__":
    main()
