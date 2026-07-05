#!/usr/bin/env python3
"""Dump the included-studies sheet and match each study to its publication files.

This gives you, in one place, everything you need to start tagging: the ID column
that keys each study, the abstract (which usually states population/design/the key
comparison — enough to tag most studies without opening the PDF), and the reference
files that belong to each study (so an agent can open the full text only when the
abstract leaves the question-fit ambiguous).

Usage:
  python inspect_studies.py <studies.xlsx> [--refs REF_DIR]
                            [--sheet NAME] [--id-col LETTER_OR_HEADER]
                            [--out study_index.json]

Outputs a console summary and, with --out, a study_index.json:
  [{ "id", "row", "fields": {header: value}, "abstract", "matched_files": [paths] }]
"""
import argparse, json, os, re, sys

ID_HEADERS = ("id", "study id", "studyid", "paper id", "record id", "refid", "ref id")
ABSTRACT_HEADERS = ("abstract", "summary")
TITLE_HEADERS = ("title",)
AUTHOR_HEADERS = ("authors", "author")


def col_letter(idx):  # 1-based -> A..
    s = ""
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def pick_id_col(headers, override):
    if override:
        ov = override.strip()
        if re.fullmatch(r"[A-Za-z]+", ov) and ov.upper() == ov and len(ov) <= 3:
            # looks like a column letter
            n = 0
            for ch in ov.upper():
                n = n * 26 + (ord(ch) - 64)
            return n
        for i, h in enumerate(headers, 1):
            if h and h.strip().lower() == ov.lower():
                return i
        sys.exit(f"--id-col '{override}' not found")
    for i, h in enumerate(headers, 1):
        if h and h.strip().lower() in ID_HEADERS:
            return i
    return 1  # fall back to first column


def find_col(headers, names):
    for i, h in enumerate(headers, 1):
        if h and h.strip().lower() in names:
            return i
    return None


def parse_ref_files(ref_dir):
    """Index reference files by (surname, year) parsed from common filename styles
    like '2018_Francis_NEJM.pdf' or 'Francis_2018_suppl.pdf'."""
    out = []
    if not ref_dir or not os.path.isdir(ref_dir):
        return out
    for fn in sorted(os.listdir(ref_dir)):
        if fn.startswith("."):
            continue
        path = os.path.join(ref_dir, fn)
        if not os.path.isfile(path):
            continue
        year = None
        m = re.search(r"(19|20)\d{2}", fn)
        if m:
            year = m.group(0)
        # surname = the alphabetic token that isn't a journal-ish all-caps abbrev
        tokens = re.split(r"[ _\-.]", fn)
        surname = None
        for tok in tokens:
            if re.fullmatch(r"[A-Za-z]{3,}", tok) and not tok.isupper() and tok.lower() not in (
                "suppl", "supplement", "sup", "appendix", "figure", "table", "pdf", "docx"):
                surname = tok.lower()
                break
        out.append({"file": fn, "path": path, "year": year, "surname": surname})
    return out


def first_author_blob(fields):
    """The first author only. Reference files are named by first author, while trial
    author lists are huge and overlap across collaborative groups (IBCSG/BIG) — matching
    the whole list cross-links unrelated papers, so restrict to the first author."""
    authors = ""
    for h, v in fields.items():
        if h and re.search(r"author", h, re.I) and v:
            authors = str(v)
            break
    if not authors:  # fall back to the title if there's no author column
        for h, v in fields.items():
            if h and re.search(r"title", h, re.I) and v:
                authors = str(v)
                break
    if ";" in authors:           # "Surname, Initials; Surname, Initials; ..."
        first = authors.split(";")[0]
    elif "," in authors:          # "Firstname Lastname, Firstname Lastname, ..."
        first = authors.split(",")[0]
    else:
        first = authors
    return first.lower()


def match_files(study, ref_index):
    """Match a study to ref files by FIRST-author surname + year within +/-1 (online-ahead
    vs print years differ). First-author matching avoids the false positives that arise from
    the large shared author lists of collaborative trial groups."""
    fields = study["fields"]
    nameblob = first_author_blob(fields)
    study_years = set(re.findall(r"(?:19|20)\d{2}", " ".join(str(v) for v in fields.values())))

    def year_ok(yr):
        if yr is None or not study_years:
            return True
        return any(abs(int(yr) - int(sy)) <= 1 for sy in study_years)

    hits = []
    for r in ref_index:
        sn, yr = r.get("surname"), r.get("year")
        if sn and sn in nameblob and year_ok(yr):
            hits.append(r["path"])
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("studies")
    ap.add_argument("--refs")
    ap.add_argument("--sheet")
    ap.add_argument("--id-col")
    ap.add_argument("--out")
    a = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(a.studies, data_only=True)
    ws = wb[a.sheet] if a.sheet else wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    id_idx = pick_id_col(headers, a.id_col)
    abs_idx = find_col(headers, ABSTRACT_HEADERS)
    title_idx = find_col(headers, TITLE_HEADERS)

    print(f"Sheet: {ws.title}  | rows: {ws.max_row}  | cols: {ws.max_column}")
    print(f"ID column: {col_letter(id_idx)} ('{headers[id_idx-1]}')"
          + (f" | Abstract: {col_letter(abs_idx)}" if abs_idx else " | Abstract: (none detected)"))
    print("\nColumns:")
    for i, h in enumerate(headers, 1):
        if h:
            print(f"  {col_letter(i)}: {h}")

    ref_index = parse_ref_files(a.refs)
    if a.refs:
        print(f"\nReference files in {a.refs}: {len(ref_index)}")

    studies = []
    print("\n=== STUDIES ===")
    for r in range(2, ws.max_row + 1):
        idv = ws.cell(row=r, column=id_idx).value
        if idv is None or not str(idv).strip():
            continue
        fields = {headers[c-1]: ws.cell(row=r, column=c).value
                  for c in range(1, ws.max_column + 1) if headers[c-1]}
        abstract = str(ws.cell(row=r, column=abs_idx).value or "") if abs_idx else ""
        title = str(ws.cell(row=r, column=title_idx).value or "") if title_idx else ""
        study = {"id": str(idv).strip(), "row": r, "fields": fields, "abstract": abstract}
        study["matched_files"] = match_files(study, ref_index) if ref_index else []
        studies.append(study)
        mf = study["matched_files"]
        flag = "" if (mf or not ref_index) else "  <-- NO FILE MATCH (check)"
        print(f"  {study['id']:>8}  {title[:60]:60}  files={len(mf)}{flag}")

    if a.out:
        with open(a.out, "w") as f:
            json.dump(studies, f, indent=2, default=str)
        print(f"\nwrote {len(studies)} studies to {a.out}")

    # warn on multi-match / no-match so the agent knows to confirm
    if ref_index:
        amb = [s for s in studies if len(s["matched_files"]) == 0]
        if amb:
            print(f"\n[!] {len(amb)} studies had NO file match — resolve before classifying: "
                  + ", ".join(s["id"] for s in amb))


if __name__ == "__main__":
    main()
