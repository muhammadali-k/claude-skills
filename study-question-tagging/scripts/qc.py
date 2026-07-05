#!/usr/bin/env python3
"""Quality-check the per-question Excel files against the source sheet and matrix.

Confirms each produced file is a faithful filtered copy: header row intact, same
column count, and the data rows are exactly the studies the matrix assigned (no
extras, none missing). Also reports coverage — which source studies landed in no
question file — because a silently dropped study is the most likely error.

Usage:
  python qc.py --source studies.xlsx --matrix matrix.json --out-dir DIR
               [--sheet NAME] [--id-col LETTER_OR_HEADER]
"""
import argparse, json, os, re, sys


def col_to_idx(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def resolve_id_col(headers, override):
    if override:
        ov = override.strip()
        if re.fullmatch(r"[A-Za-z]{1,3}", ov) and ov.upper() == ov:
            return col_to_idx(ov)
        for i, h in enumerate(headers, 1):
            if h and str(h).strip().lower() == ov.lower():
                return i
    for i, h in enumerate(headers, 1):
        if h and str(h).strip().lower() in ("id", "study id", "paper id", "record id"):
            return i
    return 1


def normalize_matrix(m):
    if isinstance(m, dict) and "questions" in m:
        qs = m["questions"]
    elif isinstance(m, dict):
        qs = [{"key": k, "study_ids": v} for k, v in m.items()]
    else:
        sys.exit("bad matrix")
    out = []
    for q in qs:
        key = str(q["key"]).strip()
        slug = q.get("slug")
        fn = q.get("filename") or (f"{re.sub(r'[^A-Za-z0-9._-]+','_',key)}_{slug}.xlsx"
                                   if slug else f"{re.sub(r'[^A-Za-z0-9._-]+','_',key)}.xlsx")
        out.append({"key": key, "filename": fn,
                    "study_ids": set(str(x).strip() for x in (q.get("study_ids") or []))})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True)
    ap.add_argument("--matrix", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--sheet")
    ap.add_argument("--id-col")
    a = ap.parse_args()

    import openpyxl
    with open(a.matrix) as f:
        questions = normalize_matrix(json.load(f))

    wb0 = openpyxl.load_workbook(a.source, data_only=True)
    sheet_name = a.sheet or wb0.sheetnames[0]
    ws0 = wb0[sheet_name]
    src_headers = [ws0.cell(row=1, column=c).value for c in range(1, ws0.max_column + 1)]
    id_idx = resolve_id_col(src_headers, a.id_col)
    src_ids = set()
    for r in range(2, ws0.max_row + 1):
        v = ws0.cell(row=r, column=id_idx).value
        if v is not None and str(v).strip():
            src_ids.add(str(v).strip())

    ok = True
    assigned = set()
    print("=== QC: per-question files ===")
    for q in questions:
        path = os.path.join(a.out_dir, q["filename"])
        if not os.path.exists(path):
            print(f"  [FAIL] {q['key']}: missing file {q['filename']}")
            ok = False
            continue
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        file_ids = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=id_idx).value
            if v is not None and str(v).strip():
                file_ids.add(str(v).strip())
        problems = []
        if headers != src_headers:
            problems.append("header/columns differ from source")
        if file_ids != q["study_ids"]:
            extra = file_ids - q["study_ids"]
            miss = q["study_ids"] - file_ids
            if extra:
                problems.append(f"unexpected rows {sorted(extra)}")
            if miss:
                problems.append(f"missing rows {sorted(miss)}")
        assigned |= file_ids
        status = "OK  " if not problems else "FAIL"
        if problems:
            ok = False
        print(f"  [{status}] {q['key']:8} rows={len(file_ids):3d}  {q['filename']}"
              + ("  :: " + "; ".join(problems) if problems else ""))

    unassigned = sorted(src_ids - assigned)
    print(f"\nCoverage: {len(assigned)}/{len(src_ids)} source studies appear in >=1 file.")
    if unassigned:
        print(f"[!] In NO file (confirm intended): {unassigned}")
    print("\nRESULT:", "PASS" if ok else "FAIL — fix the issues above")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
