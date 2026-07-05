#!/usr/bin/env python3
"""Parse an i-table template (and optional example) into machine-readable specs.

Outputs (to --out-dir):
  column_schema.json     every column -> {col, idx, header_path, group, variable, category, arm, is_meta}
  field_guide.md         human-readable column -> meaning map (grouped), handed to extractor agents
  col_types.json         (only with --example) each column -> NUMBER | LETTER | TEXT
  calibration_<id>.json  (only with --example) one example row's col->value, for agent calibration

Handles .xlsx templates with multi-row merged headers, and flat .csv/.tsv column lists.
Header block and data-start row are auto-detected (override with --header-rows / --data-start).
"""
import argparse, json, os, re, csv

ARM_RE = re.compile(r"(Treat?ment\s*Arm\s*\d|Treat?ment\s*\d|Control|Comparator|Placebo|Arm\s*\d|Group\s*\d)", re.I)

def arm_role(label):
    if label is None:
        return None
    s = str(label)
    if re.search(r"\b(control|comparator|placebo)\b", s, re.I):
        return "control"
    if ARM_RE.search(s):
        m = re.search(r"(\d)", s)
        if m:
            return f"treatment_{m.group(1)}"
    return None

def is_number(v):
    try:
        float(str(v).replace(",", "")); return True
    except Exception:
        return False

LETTERS = re.compile(r"^[A-Za-z ]+$")

def parse_xlsx(path, header_rows, data_start, example_path, out_dir):
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    anchor = {}
    for m in ws.merged_cells.ranges:
        a = (m.min_row, m.min_col)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                anchor[(r, c)] = a
    def cv(r, c):
        rr, cc = anchor.get((r, c), (r, c))
        return ws.cell(row=rr, column=cc).value
    if header_rows is None:
        hr = 1
        for m in ws.merged_cells.ranges:
            if m.min_row == 1:
                hr = max(hr, m.max_row)
        header_rows = hr
    if data_start is None:
        data_start = header_rows + 1
    schema = []
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        levels = []
        for r in range(1, header_rows + 1):
            v = cv(r, c)
            s = re.sub(r"\s*\(ID:.*?\)", "", str(v)).strip() if v is not None else ""
            if s and (not levels or levels[-1] != s):
                levels.append(s)
        if not levels:
            continue
        group = levels[0]
        variable = levels[1] if len(levels) > 1 else levels[0]
        arm = arm_role(levels[-1]) if len(levels) > 1 else None
        arm_label = levels[-1] if arm else None
        middle = levels[2:-1] if arm else levels[2:]
        category = " / ".join(middle) if middle else None
        is_meta = len(levels) == 1
        schema.append({
            "col": letter, "idx": c, "header_path": " / ".join(levels),
            "group": group, "variable": variable, "category": category,
            "arm": arm, "arm_label": arm_label, "is_meta": is_meta,
        })
    write_schema(schema, out_dir)

    if example_path:
        exwb = openpyxl.load_workbook(example_path, data_only=True)
        exws = exwb.worksheets[0]
        ex_hr = 1
        for m in exws.merged_cells.ranges:
            if m.min_row == 1:
                ex_hr = max(ex_hr, m.max_row)
        rows = list(range(ex_hr + 1, exws.max_row + 1))
        types = {}
        for s in schema:
            c = s["idx"]
            vals = [exws.cell(row=r, column=c).value for r in rows]
            nonempty = [str(v).strip() for v in vals if v is not None and str(v).strip() != ""]
            na = [v for v in nonempty if v.upper() == "NA"]
            real = [v for v in nonempty if v.upper() != "NA"]
            nums = [v for v in real if is_number(v)]
            lett = [v for v in real if LETTERS.match(v)]
            if real and len(nums) == len(real) and len(na) == 0:
                t = "NUMBER"
            elif real and len(lett) == len(real):
                t = "LETTER"
            else:
                t = "TEXT"
            types[s["col"]] = t
        json.dump(types, open(os.path.join(out_dir, "col_types.json"), "w"), indent=1)
        if rows:
            r0 = rows[0]
            gold = {get_column_letter(c): str(exws.cell(row=r0, column=c).value)
                    for c in range(1, exws.max_column + 1)
                    if exws.cell(row=r0, column=c).value not in (None, "")}
            cid = gold.get("A", "row")
            json.dump({"row": r0, "values": gold},
                      open(os.path.join(out_dir, f"calibration_{cid}.json"), "w"), indent=1)
        print(f"Derived col_types.json ({sum(1 for v in types.values() if v=='NUMBER')} NUMBER, "
              f"{sum(1 for v in types.values() if v=='LETTER')} LETTER) and calibration.")
    print(f"Parsed {len(schema)} columns (header_rows={header_rows}, data_start={data_start}).")

def parse_csv(path, out_dir):
    delim = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, newline="") as f:
        headers = next(csv.reader(f, delimiter=delim))
    from openpyxl.utils import get_column_letter
    schema = []
    for i, h in enumerate(headers, start=1):
        schema.append({"col": get_column_letter(i), "idx": i, "header_path": h.strip(),
                       "group": h.strip(), "variable": h.strip(), "category": None,
                       "arm": arm_role(h), "arm_label": (h.strip() if arm_role(h) else None),
                       "is_meta": arm_role(h) is None})
    write_schema(schema, out_dir)
    print(f"Parsed {len(schema)} columns from flat {os.path.basename(path)} (no data types without --example).")

def write_schema(schema, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    json.dump(schema, open(os.path.join(out_dir, "column_schema.json"), "w"), indent=1)
    from collections import OrderedDict
    guide = OrderedDict()
    for col in schema:
        key = (col["group"], col["variable"], col["category"])
        guide.setdefault(key, []).append((col["arm"] or "(single)", col["col"]))
    lines = ["# Field guide (group / variable / category -> arm:column)\n"]
    cur = None
    for (group, variable, category), arms in guide.items():
        if group != cur:
            lines.append(f"\n## {group}"); cur = group
        cat = f" [{category}]" if category else ""
        vv = "" if variable == group else f"{variable}"
        lines.append(f"- {vv}{cat}  ->  " + ", ".join(f"{a}:{c}" for a, c in arms))
    open(os.path.join(out_dir, "field_guide.md"), "w").write("\n".join(lines))

def main():
    ap = argparse.ArgumentParser(description="Parse i-table template into schema + field guide (+ types/calibration).")
    ap.add_argument("--template", required=True, help="Input sheet (.xlsx/.csv/.tsv) defining the columns")
    ap.add_argument("--example", help="Optional example filled .xlsx (derives data types + calibration)")
    ap.add_argument("--out-dir", default="_work")
    ap.add_argument("--header-rows", type=int, default=None, help="Override auto-detected header row count")
    ap.add_argument("--data-start", type=int, default=None, help="Override auto-detected first data row")
    a = ap.parse_args()
    os.makedirs(a.out_dir, exist_ok=True)
    if a.template.lower().endswith((".csv", ".tsv")):
        parse_csv(a.template, a.out_dir)
        if a.example:
            print("Note: --example type derivation requires an .xlsx template; skipped for flat input.")
    else:
        parse_xlsx(a.template, a.header_rows, a.data_start, a.example, a.out_dir)
    print(f"Wrote column_schema.json + field_guide.md to {a.out_dir}/")

if __name__ == "__main__":
    main()
