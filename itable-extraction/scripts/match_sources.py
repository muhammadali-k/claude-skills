#!/usr/bin/env python3
"""Match source publication files to the pre-seeded rows of an i-table input sheet.

Matches by first-author surname + year (and DOI/PMID if the filename carries it), groups each main text
with its supplements, and FLAGS ambiguous rows (0 or >1 candidate main files) for human review.

Output: sources_map.json
  { "<row_id>": {"row": <r>, "author": "...", "year": "...", "main": "<path or null>",
                 "supplements": [...], "all": [...], "candidates": [...], "status": "ok|REVIEW"} }
The match is a STARTING POINT — always review flagged rows before extracting; a wrong file->row poisons
the whole row.
"""
import argparse, json, os, re, glob

SUPPL_RE = re.compile(r"(suppl|supp|_sup|appendix|supplementary|_fig|figure|table_s)", re.I)

def norm_tokens(s):
    return set(re.findall(r"[a-z]+", str(s).lower()))

def surname_tokens(author):
    """Return the alpha tokens of the first author's surname (handles hyphenated names like
    'Tjan-Heijnen' -> {'tjan','heijnen'}), keeping tokens >=3 chars to avoid weak matches."""
    if not author:
        return set()
    first = re.split(r"[;]", str(author).strip())[0]
    surname = re.split(r"[,]", first)[0].strip() if "," in first else (first.split()[-1] if first.split() else "")
    return {t for t in re.findall(r"[a-z]+", surname.lower()) if len(t) >= 3}

def year_of(*vals):
    for v in vals:
        if v is None:
            continue
        m = re.search(r"(19|20)\d{2}", str(v))
        if m:
            return m.group(0)
    return None

def load_rows(sheet):
    import openpyxl
    wb = openpyxl.load_workbook(sheet, data_only=True)
    ws = wb.worksheets[0]
    hr = 1
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    def find(*names):
        for nm, c in hdr.items():
            if any(k.lower() in str(nm).lower() for k in names):
                return c
        return None
    c_id = find("paper id", "id", "study") or 1
    c_auth = find("author") or 2
    c_title = find("title") or 3
    c_date = find("date", "year") or 4
    c_doi = find("doi", "publication id", "pmid", "nct") or 6
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        rid = ws.cell(row=r, column=c_id).value
        if rid in (None, ""):
            continue
        rows.append({
            "row": r, "id": str(rid).strip(),
            "author": ws.cell(row=r, column=c_auth).value,
            "title": ws.cell(row=r, column=c_title).value,
            "year": year_of(ws.cell(row=r, column=c_date).value),
            "ident": str(ws.cell(row=r, column=c_doi).value or "").lower(),
        })
    return rows

def main():
    ap = argparse.ArgumentParser(description="Auto-match source files to i-table rows by metadata.")
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--sources-dir", required=True)
    ap.add_argument("--out", default="_work/sources_map.json")
    ap.add_argument("--exts", default=".pdf,.docx,.doc,.txt")
    a = ap.parse_args()
    exts = tuple(e.strip() for e in a.exts.split(","))
    files = [p for p in glob.glob(os.path.join(a.sources_dir, "**", "*"), recursive=True)
             if p.lower().endswith(exts)]
    rows = load_rows(a.sheet)

    def stem(path):
        b = os.path.splitext(os.path.basename(path))[0].lower()
        b = re.sub(r"[ _\-]*(suppl?|supplementary|sup|appendix|fig(ure)?\d*|table[_ ]?s?\d*|s\d+)[ _\-]*\d*$", "", b)
        return re.sub(r"[ _\-]+$", "", b)

    out = {}
    for row in rows:
        stoks = surname_tokens(row["author"])
        yr = row["year"]
        ident = row["ident"]
        cands = []
        for p in files:
            base = os.path.basename(p)
            toks = norm_tokens(base)
            frag = re.sub(r"[^a-z0-9]", "", ident)[-6:] if ident else ""
            ident_hit = bool(frag) and frag in re.sub(r"[^a-z0-9]", "", base.lower())
            surname_hit = bool(stoks and (stoks & toks))
            if not (surname_hit or ident_hit):
                continue
            score = (2 if surname_hit else 0) + (1 if (yr and yr in base) else 0) + (3 if ident_hit else 0)
            cands.append((score, p))
        cands.sort(reverse=True, key=lambda x: x[0])
        paths = [p for _, p in cands]
        mains = [p for p in paths if not SUPPL_RE.search(os.path.basename(p))]
        best_main = mains[0] if mains else None
        if best_main:
            ms = stem(best_main)
            suppls = [p for p in paths if SUPPL_RE.search(os.path.basename(p)) and stem(p) == ms]
        else:
            suppls = [p for p in paths if SUPPL_RE.search(os.path.basename(p))]
        top = cands[0][0] if cands else 0
        distinct_top_mains = [p for sc, p in cands if sc == top and not SUPPL_RE.search(os.path.basename(p))]
        status = "ok" if len(distinct_top_mains) == 1 else "REVIEW"
        out[row["id"]] = {
            "row": row["row"], "author": str(row["author"] or "")[:60], "year": yr,
            "main": best_main, "supplements": suppls, "all": paths,
            "candidates": [os.path.basename(p) for p in paths], "status": status,
        }
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    json.dump(out, open(a.out, "w"), indent=1)
    n_review = sum(1 for v in out.values() if v["status"] == "REVIEW")
    print(f"Matched {len(out)} rows -> {a.out}. {n_review} need REVIEW (0 or >1 main candidate):")
    for rid, v in out.items():
        if v["status"] == "REVIEW":
            print(f"  ! {rid} ({v['author'][:24]} {v['year']}): candidates={v['candidates']}")
    print("Review flagged rows with the user before extracting.")

if __name__ == "__main__":
    main()
