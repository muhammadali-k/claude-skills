#!/usr/bin/env python3
"""Write reconciled extraction results into the i-table, fill missing cells per column type, and emit a
provenance workbook + long-format CSV.

  python assemble.py --results _work/extraction_results.json --template input.xlsx \
      --schema _work/column_schema.json [--col-types _work/col_types.json] \
      (--in-place | --out filled.xlsx) [--number-missing blank|zero] [--text-missing NA]

extraction_results.json: {"studies":[{"id"|"paper_id","arm_mapping","cells":[{col,value,source}],"flags",
"verify_note","reconcile_changes"}]}. Missing cells are filled with the marker the column's TYPE allows:
text/letter -> --text-missing (default 'NA'); number -> blank (or 0 with --number-missing zero). Without
--col-types, all missing data cells get the text marker and a warning is printed (run validate_types after).
"""
import argparse, json, os, shutil, csv, datetime

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--results", required=True)
    ap.add_argument("--template", required=True, help="Input sheet to fill (xlsx)")
    ap.add_argument("--schema", required=True)
    ap.add_argument("--col-types")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--out")
    ap.add_argument("--id-col", default="A")
    ap.add_argument("--number-missing", choices=["blank", "zero"], default="blank")
    ap.add_argument("--text-missing", default="NA")
    a = ap.parse_args()
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter

    schema = {c["col"]: c for c in json.load(open(a.schema))}
    data_cols = [c["col"] for c in schema.values() if not c.get("is_meta")]
    types = json.load(open(a.col_types)) if a.col_types else None
    studies = json.load(open(a.results))["studies"]

    target = a.template if a.in_place else a.out
    if not target:
        ap.error("Provide --in-place or --out")
    if not a.in_place:
        shutil.copy2(a.template, a.out)
    stamp = datetime.date.today().isoformat()
    backup = os.path.splitext(target)[0] + f"_BACKUP_{stamp}.xlsx"
    if not os.path.exists(backup):
        shutil.copy2(a.template if a.in_place else a.out, backup)

    wb = openpyxl.load_workbook(target)
    ws = wb.worksheets[0]
    id_ci = column_index_from_string(a.id_col)
    hr = 1
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)
    id_row = {}
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=id_ci).value
        if v not in (None, ""):
            id_row[str(v).strip()] = r

    written = {}; skipped = []
    for st in studies:
        sid = str(st.get("id") or st.get("paper_id")).strip()
        if sid not in id_row:
            skipped.append(sid); continue
        r = id_row[sid]; n = 0
        for cell in st.get("cells", []):
            col = str(cell["col"]).strip()
            try:
                ci = column_index_from_string(col)
            except Exception:
                continue
            if col not in schema or schema[col].get("is_meta"):
                continue
            val = cell.get("value")
            if val is None or str(val).strip() == "":
                continue
            cval = str(val).strip()
            if types and types.get(col) == "NUMBER":
                try:
                    cval = int(cval) if float(cval) == int(float(cval)) else float(cval)
                except Exception:
                    pass
            ws.cell(row=r, column=ci).value = cval
            n += 1
        written[sid] = n

    na_filled = {}
    for sid, r in id_row.items():
        k = 0
        for col in data_cols:
            ci = column_index_from_string(col)
            cell = ws.cell(row=r, column=ci)
            if cell.value not in (None, ""):
                continue
            t = types.get(col) if types else None
            if t == "NUMBER":
                if a.number_missing == "zero":
                    cell.value = 0; k += 1
            else:
                cell.value = a.text_missing; k += 1
        na_filled[sid] = k

    wb.save(target)
    if not types:
        print("WARNING: no --col-types; filled all missing data cells with the text marker. Run "
              "validate_types.py before upload to fix any numeric/letter columns.")
    print(f"Saved: {target}\n  cells written: {written}\n  missing-filled: {na_filled}")
    if skipped:
        print("  WARNING ids not found in sheet:", skipped)

    base = os.path.splitext(target)[0]
    prov = base + "_provenance.xlsx"
    pwb = openpyxl.Workbook(); ps = pwb.active; ps.title = "Provenance"
    ps.append(["Study ID", "Col", "Header path", "Arm", "Value", "Source"])
    for st in studies:
        for cell in st.get("cells", []):
            sc = schema.get(str(cell["col"]).strip(), {})
            ps.append([str(st.get("id") or st.get("paper_id")), cell["col"], sc.get("header_path", ""), sc.get("arm", ""),
                       cell.get("value", ""), cell.get("source", "")])
    fs = pwb.create_sheet("Arm_Mapping_and_Flags")
    fs.append(["Study ID", "N arms", "Arm mapping", "Flags", "Verifier note", "Reconcile changes"])
    for st in studies:
        fs.append([str(st.get("id") or st.get("paper_id")), st.get("n_arms", ""),
                   json.dumps(st.get("arm_mapping", {}), ensure_ascii=False),
                   " | ".join(st.get("flags", []) or []), st.get("verify_note", "") or "",
                   json.dumps(st.get("reconcile_changes", []), ensure_ascii=False)])
    ss = pwb.create_sheet("Summary")
    ss.append(["Study ID", "Cells written", "Missing-filled"])
    for st in studies:
        sid = str(st.get("id") or st.get("paper_id")).strip()
        ss.append([sid, written.get(sid, 0), na_filled.get(sid, "")])
    pwb.save(prov)

    long_csv = base + "_long.csv"
    with open(long_csv, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["study_id", "col", "header_path", "arm", "value", "source"])
        for st in studies:
            for cell in st.get("cells", []):
                sc = schema.get(str(cell["col"]).strip(), {})
                w.writerow([str(st.get("id") or st.get("paper_id")), cell["col"], sc.get("header_path", ""), sc.get("arm", ""),
                            cell.get("value", ""), cell.get("source", "")])
    print(f"  provenance -> {prov}\n  long CSV   -> {long_csv}\n  backup     -> {backup}")

if __name__ == "__main__":
    main()
