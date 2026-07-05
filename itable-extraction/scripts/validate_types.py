#!/usr/bin/env python3
"""Validate (and optionally fix) a filled i-table against per-column data types so it imports cleanly.

  python validate_types.py --sheet filled.xlsx (--col-types _work/col_types.json | --example ex.xlsx) \
      [--fix] [--out out.xlsx] [--number-missing blank|zero]

Rules (see references/data_types.md):
  NUMBER column -> every non-empty cell must be a number; 'NA'/text -> blank (or 0 with --number-missing zero)
  LETTER column -> letters+spaces only ('NA' allowed); offending cells sanitized (+->plus, ->->then, digits
                    spelled, symbols dropped)
  TEXT column   -> anything allowed
Without --fix it only reports. Writes a safety copy before fixing. Leaves a Validation_Fixes log next to
the sheet (csv). Note: filling real identifiers (e.g. a missing PubMed ID) is a manual step, not done here.
"""
import argparse, json, os, re, shutil, csv

LETTERS = re.compile(r"^[A-Za-z ]+$")
NUMWORDS = {"0": "", "1": "one", "2": "two", "3": "three", "4": "four", "5": "five",
            "6": "six", "7": "seven", "8": "eight", "9": "nine", "10": "ten", "11": "eleven", "12": "twelve"}

def is_num(v):
    try:
        float(str(v).replace(",", "")); return True
    except Exception:
        return False

def to_letters(s):
    s = s.replace("+", " plus ").replace("->", " then ").replace("→", " then ").replace("/", " or ").replace("&", " and ")
    s = re.sub(r"\b(\d+)\b", lambda m: NUMWORDS.get(m.group(1), ""), s)
    s = re.sub(r"[^A-Za-z ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def derive_types(example):
    import openpyxl
    from openpyxl.utils import get_column_letter
    ws = openpyxl.load_workbook(example, data_only=True).worksheets[0]
    hr = 1
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)
    types = {}
    for c in range(1, ws.max_column + 1):
        vals = [ws.cell(row=r, column=c).value for r in range(hr + 1, ws.max_row + 1)]
        ne = [str(v).strip() for v in vals if v is not None and str(v).strip() != ""]
        na = [v for v in ne if v.upper() == "NA"]
        real = [v for v in ne if v.upper() != "NA"]
        nums = [v for v in real if is_num(v)]
        lett = [v for v in real if LETTERS.match(v)]
        if real and len(nums) == len(real) and not na:
            t = "NUMBER"
        elif real and len(lett) == len(real):
            t = "LETTER"
        else:
            t = "TEXT"
        types[get_column_letter(c)] = t
    return types

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--col-types")
    ap.add_argument("--example")
    ap.add_argument("--fix", action="store_true")
    ap.add_argument("--out")
    ap.add_argument("--number-missing", choices=["blank", "zero"], default="blank")
    a = ap.parse_args()
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string

    if a.col_types:
        types = json.load(open(a.col_types))
    elif a.example:
        types = derive_types(a.example)
    else:
        ap.error("Provide --col-types or --example")
    NUM = {c for c, t in types.items() if t == "NUMBER"}
    LET = {c for c, t in types.items() if t == "LETTER"}

    target = a.sheet if not a.out else a.out
    if a.out:
        shutil.copy2(a.sheet, a.out)
    if a.fix:
        safety = os.path.splitext(target)[0] + "_PREVALIDATIONFIX.xlsx"
        shutil.copy2(target, safety)
    wb = openpyxl.load_workbook(target)
    ws = wb.worksheets[0]
    hr = 1
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)

    num_bad = []; let_bad = []; fixes = []
    for r in range(hr + 1, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        for c in range(1, ws.max_column + 1):
            L = get_column_letter(c); cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None or str(v).strip() == "":
                continue
            s = str(v).strip()
            if L in NUM and not is_num(s):
                num_bad.append((r, L, s))
                if a.fix:
                    cell.value = 0 if a.number_missing == "zero" else None
                    fixes.append((r, L, s, cell.value, "NUMBER col: non-number -> missing"))
            elif L in LET and s.upper() != "NA" and not LETTERS.match(s):
                let_bad.append((r, L, s))
                if a.fix:
                    new = to_letters(s)
                    cell.value = new
                    fixes.append((r, L, s, new, "LETTER col: sanitized to letters+spaces"))

    print(f"NUMBER violations: {len(num_bad)} | LETTER violations: {len(let_bad)}")
    for x in (num_bad + let_bad)[:30]:
        print("   !", x)
    if a.fix:
        wb.save(target)
        log = os.path.splitext(target)[0] + "_validation_fixes.csv"
        with open(log, "w", newline="") as f:
            w = csv.writer(f); w.writerow(["row", "col", "old", "new", "reason"])
            w.writerows(fixes)
        print(f"Applied {len(fixes)} fixes -> {target} (safety copy + {os.path.basename(log)}).")
        print("Re-run without --fix to confirm 0 violations.")
    else:
        print("OK — no violations." if not num_bad and not let_bad else "Run with --fix to repair.")

if __name__ == "__main__":
    main()
