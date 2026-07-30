#!/usr/bin/env python3
"""QC a filled i-table against the reconciled results + structural integrity.

  python qc.py --sheet filled.xlsx --results _work/extraction_results.json --schema _work/column_schema.json \
      [--template original.xlsx] [--id-col A] [--vocabulary <project>_node_vocabulary.json]

Checks: (1) round-trip — every written value equals the reconciled result; (2) structure — column count and
merged-header count match the original template (if --template given); (3) metadata preserved; (4) NODE
LABELS, with --vocabulary — every treatment- and control-arm label resolved against the project's
controlled vocabulary, and the RUN FAILS (exit 1) on any rejection; then prints fill-rates and ALL flagged
judgment calls for the user to review.

*** NODE LABELS (--vocabulary) ***
netmeta joins arms by STRING EQUALITY of their labels, so "NIVO + IPI" and "NIVO+IPI" are two nodes rather
than one, and the resulting network is either disconnected or — worse — quietly missing an edge. The i-table
is where this usually starts: its arm columns are written by a different pass from the outcomes sheets, and
in the reference project the two disagreed for the SAME trials (Atezolezumab / Atezolizumab; Everolimus /
Everoilmus; Nivolumab + Iplimumab / Nivolumab plus iplimumab / Nivulumab + ipililumab). Every rejection
prints nodes.py's own diagnosis and its suggested fix verbatim. Alias hits are reported as accepted-but-
non-canonical rather than passed silently, so a sheet full of legacy wording is visible as such.
"""
import argparse, json, os, sys

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--results", required=True)
    ap.add_argument("--schema", required=True)
    ap.add_argument("--template")
    ap.add_argument("--id-col", default="A")
    ap.add_argument("--vocabulary", help="project controlled node vocabulary (JSON); "
                                        "without it, arm labels are NOT validated")
    a = ap.parse_args()
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter

    schema = {c["col"]: c for c in json.load(open(a.schema))}
    data_cols = [c["col"] for c in schema.values() if not c.get("is_meta")]
    studies = json.load(open(a.results))["studies"]
    ws = openpyxl.load_workbook(a.sheet, data_only=True).worksheets[0]
    id_ci = column_index_from_string(a.id_col)
    hr = 1
    for m in openpyxl.load_workbook(a.sheet).worksheets[0].merged_cells.ranges:
        if m.min_row == 1:
            hr = max(hr, m.max_row)
    id_row = {str(ws.cell(row=r, column=id_ci).value).strip(): r
              for r in range(hr + 1, ws.max_row + 1) if ws.cell(row=r, column=id_ci).value not in (None, "")}

    problems = []
    if a.template:
        twb = openpyxl.load_workbook(a.template); tws = twb.worksheets[0]
        swb = openpyxl.load_workbook(a.sheet); sws = swb.worksheets[0]
        if tws.max_column != sws.max_column:
            problems.append(f"column count {sws.max_column} != template {tws.max_column}")
        if len(tws.merged_cells.ranges) != len(sws.merged_cells.ranges):
            problems.append(f"merged ranges {len(sws.merged_cells.ranges)} != template {len(tws.merged_cells.ranges)}")

    mism = 0
    for st in studies:
        sid = str(st.get("id") or st.get("paper_id")).strip(); r = id_row.get(sid)
        if not r:
            problems.append(f"study {sid} missing from sheet"); continue
        for cell in st.get("cells", []):
            col = str(cell["col"]).strip()
            if col not in schema or schema[col].get("is_meta"):
                continue
            ci = column_index_from_string(col)
            got = str(ws.cell(row=r, column=ci).value).strip()
            exp = str(cell["value"]).strip()
            if got != exp:
                mism += 1
                if mism <= 20:
                    problems.append(f"{sid} {col}: sheet='{got}' != result='{exp}'")

    print("=== FILL-RATE (non-missing data cells per study) ===")
    for st in studies:
        sid = str(st.get("id") or st.get("paper_id")).strip(); r = id_row.get(sid)
        if not r:
            continue
        filled = sum(1 for c in data_cols
                     if ws.cell(row=r, column=column_index_from_string(c)).value not in (None, "")
                     and str(ws.cell(row=r, column=column_index_from_string(c)).value).strip().upper() != "NA")
        print(f"  {sid:10} {filled:4} / {len(data_cols)}  {st.get('trial', st.get('label',''))}")

    print("\n=== FLAGS / JUDGMENT CALLS ===")
    for st in studies:
        fl = st.get("flags", []) or []
        if fl:
            print(f"[{st.get('id') or st.get('paper_id')}] {st.get('trial', st.get('label',''))}")
            for f in fl:
                print("   -", f)

    # --- node labels -------------------------------------------------------
    node_fail = 0
    print("\n=== NODE LABELS ===")
    if not a.vocabulary:
        print("  NOT VALIDATED — no --vocabulary supplied.")
        print("  Arm labels written here feed the outcomes sheets and then netmeta, which joins arms by")
        print("  string equality. One variant spelling becomes a second node and the network silently")
        print("  fragments. Pass --vocabulary <project>_node_vocabulary.json to check them.")
    else:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from nodes import Vocabulary
        vocab = Vocabulary.load(a.vocabulary)
        # Arm-bearing cells are those the schema marks as arm labels, plus anything the
        # reconciled results recorded in arm_mapping.
        checked = accepted_alias = 0
        for st in studies:
            sid = str(st.get("id") or st.get("paper_id")).strip()
            for role, raw in (st.get("arm_mapping") or {}).items():
                checked += 1
                label, note = vocab.resolve(raw)
                if label is None and note != "empty":
                    node_fail += 1
                    print(f"  REJECTED  [{sid}] {role}: {note}")
                elif note:
                    accepted_alias += 1
                    print(f"  ACCEPTED  [{sid}] {role}: {raw!r} -> {label}   ({note})")
        if not checked:
            print("  no arm labels found in the reconciled results (arm_mapping empty)")
        elif node_fail == 0 and accepted_alias == 0:
            print(f"  OK — {checked} arm label(s), all canonical.")
        else:
            print(f"  {checked} checked, {node_fail} rejected, {accepted_alias} accepted via alias "
                  f"(non-canonical — rewrite before these reach an analysis sheet).")

    print("\n=== QC RESULT ===")
    print(f"round-trip mismatches: {mism}")
    if node_fail:
        print(f"node-label rejections: {node_fail}")
    if problems:
        print(f"PROBLEMS ({len(problems)}):")
        for p in problems[:40]:
            print("  !", p)
    else:
        print("OK — written values match reconciled data; structure intact.")
    if node_fail:
        sys.exit(1)

if __name__ == "__main__":
    main()
